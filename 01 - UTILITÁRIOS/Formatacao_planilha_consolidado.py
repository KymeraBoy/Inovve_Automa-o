import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import re


# ============================================================
# CONFIGURAÇÕES
# ============================================================

pasta = Path(__file__).parent

arquivo_entrada = pasta / "consolidado.xlsx"
arquivo_saida = pasta / "consolidado_formatado.xlsx"
arquivo_entrada_valores = pasta / "consolidado_valores.xlsx"


# ============================================================
# MAPA DOS MESES EM PORTUGUÊS
# ============================================================

# LAYOUT 1
# meses = {
#     "jan": 1,
#     "fev": 2,
#     "mar": 3,
#     "abr": 4,
#     "mai": 5,
#     "jun": 6,
#     "jul": 7,
#     "ago": 8,
#     "set": 9,
#     "out": 10,
#     "nov": 11,
#     "dez": 12
# }

# LAYOUT 3
meses = {
    "JANEIRO": 1,
    "FEVEREIRO": 2,
    "MARÇO": 3,
    "ABRIL": 4,
    "MAIO": 5,
    "JUNHO": 6,
    "JULHO": 7,
    "AGOSTO": 8,
    "SETEMBRO": 9,
    "OUTUBRO": 10,
    "NOVEMBRO": 11,
    "DEZEMBRO": 12
}


# ============================================================
# SALVA A PLANILHA COM OS VALORES CALCULADOS
# ============================================================

def salvar_planilha_com_valores(arquivo_origem, arquivo_destino):

    wb_formulas = load_workbook(
        arquivo_origem,
        data_only=False
    )

    wb_valores = load_workbook(
        arquivo_origem,
        data_only=True
    )

    formulas_sem_valor = []

    for ws_formulas in wb_formulas.worksheets:

        ws_valores = wb_valores[ws_formulas.title]

        for linha in ws_formulas.iter_rows():

            for celula in linha:

                if celula.data_type == "f":

                    celula_valor = ws_valores[celula.coordinate].value
                    celula.value = celula_valor

                    if celula_valor is None:
                        formulas_sem_valor.append(
                            f"{ws_formulas.title}!{celula.coordinate}"
                        )

    wb_formulas.save(arquivo_destino)
    wb_formulas.close()
    wb_valores.close()

    if formulas_sem_valor:
        print(
            "Aviso: estas fórmulas não possuem valor calculado salvo "
            f"no arquivo: {', '.join(formulas_sem_valor)}"
        )


# ============================================================
# FUNÇÃO PARA CONVERTER MÊS/ANO
# ============================================================

def converter_mes_ano(valor):

    if valor is None:
        return None

    if isinstance(valor, datetime):
        return valor

    texto = str(valor).strip().lower()
    texto = texto.replace(" ", "")

    # Exemplo: ago-18 / ago/18 / ago-2018
    padrao = re.match(
        r"([a-zç]+)[-/](\d{2,4})",
        texto
    )

    if padrao:

        mes_texto = padrao.group(1)
        ano_texto = padrao.group(2)

        if mes_texto in meses:

            mes = meses[mes_texto]
            ano = int(ano_texto)

            if ano < 100:
                ano += 2000

            return datetime(
                ano,
                mes,
                1
            )

        
    # LAYOUT 1
    # Exemplo: 08/2018
    # padrao = re.match(
    #     r"(\d{1,2})[/\-](\d{4})",
    #     texto
    # )

    # LAYOUT 3
    # Exemplo: SETEMBRO/2022
    padrao = re.match(
    r"(JANEIRO|FEVEREIRO|MARÇO|ABRIL|MAIO|JUNHO|JULHO|AGOSTO|SETEMBRO|OUTUBRO|NOVEMBRO|DEZEMBRO)/(\d{4})",
    texto.upper()
    )


    if padrao:
        # LAYOUT 1
        # mes = int(padrao.group(1))
        # ano = int(padrao.group(2))

        # LAYOUT 3
        mes = meses.get(padrao.group(1))
        ano = int(padrao.group(2))

        if 1 <= mes <= 12:

            return datetime(
                ano,
                mes,
                1
            )

    return None


# ============================================================
# FUNÇÃO PARA CONVERTER NÚMEROS
# ============================================================

def converter_numero(valor):

    if valor is None:
        return None

    # Se já é número
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()

    if not texto:
        return None

    # Formato brasileiro
    # Exemplo: 64.579,051
    if "," in texto:

        texto = texto.replace(".", "")
        texto = texto.replace(",", ".")

    try:

        return float(texto)

    except:

        return None


# ============================================================
# VERIFICA ARQUIVO
# ============================================================

if not arquivo_entrada.exists():

    print(
        f"Arquivo não encontrado: "
        f"{arquivo_entrada}"
    )

    input("Pressione ENTER para sair...")
    exit()


# ============================================================
# 1. FORMATAÇÃO INICIAL
# ============================================================

print()
print("==============================================")
print("ETAPA 1 - FORMATAÇÃO")
print("==============================================")
print()


try:

    salvar_planilha_com_valores(
        arquivo_entrada,
        arquivo_entrada_valores
    )

    planilhas = pd.read_excel(
        arquivo_entrada_valores,
        sheet_name=None,
        header=None
    )

    with pd.ExcelWriter(
        arquivo_saida,
        engine="openpyxl"
    ) as writer:

        for nome_aba, df in planilhas.items():

            print(
                f"Formatando: {nome_aba}"
            )
            # LAYOUT 1
            # df = df.iloc[7:]
            # df = df.iloc[:, 2:]
            # LAYOUT 3
            df = df.iloc[6:]
            df = df.iloc[:, 4:]

            # Reinicia índice
            df = df.reset_index(drop=True)

            # Limita nome da aba a 31 caracteres
            nome_aba_excel = nome_aba[:31]

            df.to_excel(
                writer,
                sheet_name=nome_aba_excel,
                index=False,
                header=False
            )

except Exception as e:

    print()
    print("ERRO:")
    print(e)

    input("Pressione ENTER para sair...")
    exit()


# ============================================================
# 2. ABRE ARQUIVO FORMATADO
# ============================================================

wb = load_workbook(
    arquivo_saida
)


# ============================================================
# 3. CRIA ABA ANALISE
# ============================================================

if "ANALISE" in wb.sheetnames:
    del wb["ANALISE"]

ws_analise = wb.create_sheet(
    "ANALISE",
    0
)


# ============================================================
# 4. TÍTULO
# ============================================================

ws_analise["A1"] = (
    "ANÁLISE DE CONSUMO E ILUMINAÇÃO"
)

ws_analise["A1"].font = Font(
    bold=True,
    size=16,
    color="FFFFFF"
)

ws_analise["A1"].fill = PatternFill(
    "solid",
    fgColor="1F4E78"
)

ws_analise["A1"].alignment = Alignment(
    horizontal="center"
)

ws_analise.merge_cells(
    "A1:E1"
)


# ============================================================
# 5. CABEÇALHO DA TABELA
# ============================================================

cabecalhos = [
    "MÊS - ANO",
    "QTD. DIAS",
    "CONSUMO TOTAL - KWH",
    "QTD LÂMPADAS",
    "CONSUMO CORRIGIDO - 30 DIAS"
]


for coluna, titulo in enumerate(
    cabecalhos,
    start=1
):

    cell = ws_analise.cell(
        row=3,
        column=coluna,
        value=titulo
    )

    cell.font = Font(
        bold=True,
        color="FFFFFF"
    )

    cell.fill = PatternFill(
        "solid",
        fgColor="4472C4"
    )

    cell.alignment = Alignment(
        horizontal="center"
    )


# ============================================================
# 6. COLETA OS DADOS DAS ABAS
# ============================================================

dados_analise = []

print()
print("==============================================")
print("ETAPA 2 - ANÁLISE DOS PERÍODOS")
print("==============================================")
print()


for nome_aba in wb.sheetnames:

    if nome_aba == "ANALISE":
        continue

    ws = wb[nome_aba]

    print(f"Aba: {nome_aba}")

    # --------------------------------------------------------
    # Valores
    # --------------------------------------------------------

    # LAYOUT 1
    # mes_ano_original = ws["B2"].value
    # qtd_dias_original = ws["B3"].value
    # consumo_original = ws["B4"].value
    # lampadas_original = ws["B6"].value
    # LAYOUT 3
    mes_ano_original = ws["B2"].value
    qtd_dias_original = ws["B3"].value
    consumo_original = ws["B8"].value
    lampadas_original = ws["B4"].value

    print(
        f"  MÊS - ANO: {mes_ano_original}"
    )

    print(
        f"  QTD. DIAS: {qtd_dias_original}"
    )

    print(
        f"  CONSUMO: {consumo_original}"
    )

    print(
        f"  LÂMPADAS: {lampadas_original}"
    )

    # --------------------------------------------------------
    # Converte mês
    # --------------------------------------------------------

    data = converter_mes_ano(
        mes_ano_original
    )

    if data is None:

        print(
            "  >>> IGNORADA: "
            "mês não reconhecido."
        )

        print()
        continue

    # --------------------------------------------------------
    # Converte dias
    # --------------------------------------------------------

    qtd_dias = converter_numero(
        qtd_dias_original
    )

    # --------------------------------------------------------
    # Converte consumo
    # --------------------------------------------------------

    consumo = converter_numero(
        consumo_original
    )

    # --------------------------------------------------------
    # Converte lâmpadas
    # --------------------------------------------------------

    lampadas = converter_numero(
        lampadas_original
    )

    # --------------------------------------------------------
    # Verifica dados necessários
    # --------------------------------------------------------

    if qtd_dias is None:

        print(
            "  >>> IGNORADA: "
            "quantidade de dias inválida."
        )

        print()
        continue

    if consumo is None:

        print(
            "  >>> IGNORADA: "
            "consumo inválido."
        )

        print()
        continue

    if qtd_dias <= 0:

        print(
            "  >>> IGNORADA: "
            "quantidade de dias <= 0."
        )

        print()
        continue

    # --------------------------------------------------------
    # CALCULA CONSUMO CORRIGIDO
    # --------------------------------------------------------
    #
    # Fórmula:
    #
    # consumo corrigido =
    # consumo real / dias do mês * 30
    #
    # --------------------------------------------------------

    consumo_corrigido = (
        consumo / qtd_dias
    ) * 30

    # --------------------------------------------------------
    # Adiciona
    # --------------------------------------------------------

    dados_analise.append({

        "data": data,

        "dias": qtd_dias,

        "consumo": consumo,

        "lampadas": lampadas,

        "consumo_corrigido": consumo_corrigido,

        "aba": nome_aba

    })

    print(
        f"  >>> OK | "
        f"Consumo corrigido: "
        f"{consumo_corrigido:,.3f} kWh"
    )

    print()


# ============================================================
# 7. ORDENA CRONOLOGICAMENTE
# ============================================================

dados_analise.sort(
    key=lambda x: x["data"]
)


# ============================================================
# 8. ESCREVE DADOS NA ABA ANALISE
# ============================================================

linha = 4

for item in dados_analise:

    # Mês/Ano
    ws_analise.cell(
        row=linha,
        column=1,
        value=item["data"]
    )

    # Dias
    ws_analise.cell(
        row=linha,
        column=2,
        value=item["dias"]
    )

    # Consumo
    ws_analise.cell(
        row=linha,
        column=3,
        value=item["consumo"]
    )

    # Lâmpadas
    ws_analise.cell(
        row=linha,
        column=4,
        value=item["lampadas"]
    )

    # Consumo corrigido
    ws_analise.cell(
        row=linha,
        column=5,
        value=item["consumo_corrigido"]
    )

    # --------------------------------------------------------
    # Formatação
    # --------------------------------------------------------

    ws_analise.cell(
        row=linha,
        column=1
    ).number_format = "mmm-yy"

    ws_analise.cell(
        row=linha,
        column=2
    ).number_format = "0"

    ws_analise.cell(
        row=linha,
        column=3
    ).number_format = '#,##0.000'

    ws_analise.cell(
        row=linha,
        column=4
    ).number_format = '#,##0'

    ws_analise.cell(
        row=linha,
        column=5
    ).number_format = '#,##0.000'

    linha += 1


# ============================================================
# 9. LARGURA DAS COLUNAS
# ============================================================

ws_analise.column_dimensions["A"].width = 18
ws_analise.column_dimensions["B"].width = 14
ws_analise.column_dimensions["C"].width = 25
ws_analise.column_dimensions["D"].width = 18
ws_analise.column_dimensions["E"].width = 32


# ============================================================
# 10. GRÁFICO - CONSUMO REAL
# ============================================================

if len(dados_analise) > 0:

    grafico_consumo = LineChart()

    grafico_consumo.title = (
        "Histórico de Consumo"
    )

    grafico_consumo.y_axis.title = (
        "Consumo (kWh)"
    )

    grafico_consumo.x_axis.title = (
        "Mês/Ano"
    )

    dados = Reference(
        ws_analise,
        min_col=3,
        min_row=3,
        max_row=3 + len(dados_analise)
    )

    categorias = Reference(
        ws_analise,
        min_col=1,
        min_row=4,
        max_row=3 + len(dados_analise)
    )

    grafico_consumo.add_data(
        dados,
        titles_from_data=True
    )

    grafico_consumo.set_categories(
        categorias
    )

    grafico_consumo.height = 10
    grafico_consumo.width = 20

    ws_analise.add_chart(
        grafico_consumo,
        "G3"
    )


# ============================================================
# 11. GRÁFICO - QUANTIDADE DE LÂMPADAS
# ============================================================

if len(dados_analise) > 0:

    grafico_lampadas = LineChart()

    grafico_lampadas.title = (
        "Histórico de Quantidade de Lâmpadas"
    )

    grafico_lampadas.y_axis.title = (
        "Quantidade"
    )

    grafico_lampadas.x_axis.title = (
        "Mês/Ano"
    )

    dados = Reference(
        ws_analise,
        min_col=4,
        min_row=3,
        max_row=3 + len(dados_analise)
    )

    categorias = Reference(
        ws_analise,
        min_col=1,
        min_row=4,
        max_row=3 + len(dados_analise)
    )

    grafico_lampadas.add_data(
        dados,
        titles_from_data=True
    )

    grafico_lampadas.set_categories(
        categorias
    )

    grafico_lampadas.height = 10
    grafico_lampadas.width = 20

    ws_analise.add_chart(
        grafico_lampadas,
        "G21"
    )


# ============================================================
# 12. GRÁFICO - CONSUMO CORRIGIDO
# ============================================================

if len(dados_analise) > 0:

    grafico_corrigido = LineChart()

    grafico_corrigido.title = (
        "Histórico de Consumo Corrigido para 30 Dias"
    )

    grafico_corrigido.y_axis.title = (
        "Consumo equivalente a 30 dias (kWh)"
    )

    grafico_corrigido.x_axis.title = (
        "Mês/Ano"
    )

    dados = Reference(
        ws_analise,
        min_col=5,
        min_row=3,
        max_row=3 + len(dados_analise)
    )

    categorias = Reference(
        ws_analise,
        min_col=1,
        min_row=4,
        max_row=3 + len(dados_analise)
    )

    grafico_corrigido.add_data(
        dados,
        titles_from_data=True
    )

    grafico_corrigido.set_categories(
        categorias
    )

    grafico_corrigido.height = 11
    grafico_corrigido.width = 22

    ws_analise.add_chart(
        grafico_corrigido,
        "G39"
    )


# ============================================================
# 13. CONGELA CABEÇALHO
# ============================================================

ws_analise.freeze_panes = "A4"


# ============================================================
# 14. SALVA
# ============================================================

wb.save(
    arquivo_saida
)


# ============================================================
# 15. RESUMO
# ============================================================

print()
print("==============================================")
print("RESULTADO DA ANÁLISE")
print("==============================================")

print(
    f"Abas encontradas: "
    f"{len(wb.sheetnames) - 1}"
)

print(
    f"Períodos analisados: "
    f"{len(dados_analise)}"
)

print()
print(
    "A análise agora contém:"
)

print(
    "  - Quantidade de dias"
)

print(
    "  - Consumo real"
)

print(
    "  - Quantidade de lâmpadas"
)

print(
    "  - Consumo corrigido para 30 dias"
)

print(
    "  - Gráfico de consumo real"
)

print(
    "  - Gráfico de quantidade de lâmpadas"
)

print(
    "  - Gráfico de consumo corrigido"
)

print()
print(
    f"Arquivo criado:"
)

print(
    arquivo_saida
)

print("==============================================")

input(
    "Pressione ENTER para sair..."
)