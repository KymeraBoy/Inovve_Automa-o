from pathlib import Path
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURAÇÕES
# ============================================================

pasta = Path(__file__).parent

arquivo_saida = pasta / "consolidado.xlsx"


# ============================================================
# ENCONTRA OS ARQUIVOS
# ============================================================

arquivos = [
    arquivo
    for arquivo in pasta.iterdir()
    if arquivo.suffix.lower() in [".xlsx", ".xls"]
    and arquivo.name != arquivo_saida.name
]


if not arquivos:

    print("Nenhum arquivo .xlsx ou .xls encontrado.")

    input("Pressione ENTER para sair...")
    exit()


# ============================================================
# CRIA ARQUIVO DE SAÍDA
# ============================================================

wb_saida = Workbook()

# Remove a aba padrão
aba_padrao = wb_saida.active
wb_saida.remove(aba_padrao)


# ============================================================
# CONTROLE DE NOMES DAS ABAS
# ============================================================

nomes_abas_usados = set()


def criar_nome_aba(nome):

    # Remove caracteres inválidos
    caracteres_invalidos = [
        "\\", "/", "*", "[", "]", ":", "?"
    ]

    for caractere in caracteres_invalidos:
        nome = nome.replace(caractere, "_")

    nome = nome[:31]

    if not nome:
        nome = "Arquivo"

    nome_original = nome
    contador = 1

    while nome in nomes_abas_usados:

        sufixo = f"_{contador}"

        nome = (
            nome_original[:31 - len(sufixo)]
            + sufixo
        )

        contador += 1

    nomes_abas_usados.add(nome)

    return nome


# ============================================================
# PROCESSA CADA ARQUIVO
# ============================================================

for arquivo in arquivos:

    print()
    print("=" * 60)
    print(f"Processando: {arquivo.name}")
    print("=" * 60)

    try:

        nome_aba = criar_nome_aba(
            arquivo.stem
        )

        # ====================================================
        # XLSX
        # ====================================================

        if arquivo.suffix.lower() == ".xlsx":

            wb_origem = load_workbook(
                arquivo,
                data_only=True
            )

            ws_saida = wb_saida.create_sheet(
                nome_aba
            )

            # -----------------------------------------------
            # Se tiver apenas uma aba
            # -----------------------------------------------

            if len(wb_origem.sheetnames) == 1:

                ws_origem = wb_origem[
                    wb_origem.sheetnames[0]
                ]

                for linha in ws_origem.iter_rows():

                    for celula in linha:

                        ws_saida.cell(
                            row=celula.row,
                            column=celula.column,
                            value=celula.value
                        )

            # -----------------------------------------------
            # Se tiver várias abas
            # -----------------------------------------------

            else:

                linha_destino = 1

                for nome_aba_origem in wb_origem.sheetnames:

                    ws_origem = wb_origem[
                        nome_aba_origem
                    ]

                    # Identificação da aba original
                    ws_saida.cell(
                        row=linha_destino,
                        column=1,
                        value=f"ABA ORIGINAL: {nome_aba_origem}"
                    )

                    linha_destino += 2

                    for linha in ws_origem.iter_rows():

                        for celula in linha:

                            ws_saida.cell(
                                row=linha_destino,
                                column=celula.column,
                                value=celula.value
                            )

                        linha_destino += 1

                    linha_destino += 2

                    wb_origem.close()

        # ====================================================
        # XLS
        # ====================================================

        else:

            # Lê todas as abas do XLS
            planilhas = pd.read_excel(
                arquivo,
                sheet_name=None,
                header=None,
                engine="xlrd"
            )

            ws_saida = wb_saida.create_sheet(
                nome_aba
            )

            linha_destino = 1

            for nome_aba_origem, df in planilhas.items():

                # Se tiver várias abas
                if len(planilhas) > 1:

                    ws_saida.cell(
                        row=linha_destino,
                        column=1,
                        value=f"ABA ORIGINAL: {nome_aba_origem}"
                    )

                    linha_destino += 2

                # -------------------------------------------
                # Copia TODOS os valores
                # -------------------------------------------

                for i, linha in df.iterrows():

                    for j, valor in enumerate(
                        linha,
                        start=1
                    ):

                        if pd.notna(valor):

                            ws_saida.cell(
                                row=linha_destino,
                                column=j,
                                value=valor
                            )

                    linha_destino += 1

                linha_destino += 2

        # ====================================================
        # VERIFICA OS DADOS IMPORTANTES
        # ====================================================

        print()
        print("Verificando dados principais...")

        for linha in range(1, 10):

            valor_a = ws_saida.cell(
                row=linha,
                column=1
            ).value

            valor_b = ws_saida.cell(
                row=linha,
                column=2
            ).value

            print(
                f"Linha {linha}: "
                f"A = {valor_a!r} | "
                f"B = {valor_b!r}"
            )

        # ====================================================
        # ALERTA SE COLUNA B ESTIVER VAZIA
        # ====================================================

        valores_b = [
            ws_saida.cell(
                row=linha,
                column=2
            ).value

            for linha in range(1, 10)
        ]

        if all(
            valor is None
            for valor in valores_b
        ):

            print()
            print(
                "⚠ ATENÇÃO: "
                "a coluna B das primeiras 9 linhas "
                "está vazia!"
            )

        else:

            print()
            print(
                "✓ Dados das duas colunas "
                "encontrados."
            )

    except Exception as erro:

        print()
        print(
            f"ERRO ao processar {arquivo.name}:"
        )

        print(erro)


# ============================================================
# SALVA
# ============================================================

wb_saida.save(
    arquivo_saida
)


# ============================================================
# FINAL
# ============================================================

print()
print("=" * 60)
print("CONSOLIDAÇÃO CONCLUÍDA")
print("=" * 60)
print()
print(
    f"Arquivo criado:\n{arquivo_saida}"
)
print()
print(
    f"Total de abas: {len(wb_saida.sheetnames)}"
)
print("=" * 60)

input("Pressione ENTER para sair...")