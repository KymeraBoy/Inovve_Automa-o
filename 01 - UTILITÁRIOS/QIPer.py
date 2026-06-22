from __future__ import annotations

import argparse
from pathlib import Path
import re
from datetime import date
import os
import sys


import pandas as pd
import pdfplumber
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from concurrent.futures import ThreadPoolExecutor, as_completed

from utils import get_base_dir, clean_cell, render_progress, normalize_string_for_filename, MONTH_MAPPINGS, NUM_TO_MONTH_ABBR


BASE_DIR = get_base_dir()
DEFAULT_INPUT_DIR = BASE_DIR
OUTPUT_DIR = BASE_DIR / "Output"
DEFAULT_WORKBOOK_NAME = "QIP_consolidado.xlsx"

_MESES_ABREV = {
	1: "JAN",
	2: "FEV",
	3: "MAR",
	4: "ABR",
	5: "MAI",
	6: "JUN",
	7: "JUL",
	8: "AGO",
	9: "SET",
	10: "OUT",
	11: "NOV",
	12: "DEZ",
}

_RE_MES_ANO_REFERENCIA = re.compile(
	r"[Mm]ês/[Aa]no\s+de\s+[Rr]eferência\s+([A-Za-zçÇãõéêíóúâôáàéèíìóòúù]+)/([0-9]{4})"
)
_RE_LOCAL = re.compile(
	r"Local\s+([A-Za-zçÇãõéêíóúâôáàéèíìóòúù\s]+)(?:\n|$)",
	re.IGNORECASE,
)
_RE_ANTIGO_HEADER = re.compile(
	r"^([A-Za-zçÇãõéêíóúâôáàéèíìóòúù\s]+)\s+([0-9]{2})\s+([0-9]{4})$"
)
_NUM_REGEX = re.compile(r"^-?\d[\d\s\.,]*$")
_MAX_WORKERS = max(1, min(8, os.cpu_count() or 1))


def _render_progress(current: int, total: int, prefix: str = "Progresso") -> None:
	if total <= 0:
		return

	bar_width = 30
	ratio = current / total
	filled = int(bar_width * ratio)
	bar = "#" * filled + "-" * (bar_width - filled)
	percent = ratio * 100
	message = f"\r{prefix}: [{bar}] {current}/{total} ({percent:5.1f}%)"
	sys.stdout.write(message)
	sys.stdout.flush()

	if current >= total:
		sys.stdout.write("\n")


def _clean_cell(value: object) -> str:
	if value is None:
		return ""
	return str(value).replace("\n", " ").strip()


def _normalize_headers(headers: list[str], total_columns: int) -> list[str]:
	normalized: list[str] = []
	used_names: dict[str, int] = {}

	for index in range(total_columns):
		raw_header = headers[index] if index < len(headers) else ""
		base_name = raw_header if raw_header else f"coluna_{index + 1}"

		count = used_names.get(base_name, 0)
		used_names[base_name] = count + 1

		if count > 0:
			normalized.append(f"{base_name}_{count + 1}")
		else:
			normalized.append(base_name)

	return normalized


def _table_to_dataframe(raw_table: list[list[object]]) -> pd.DataFrame:
	cleaned_rows: list[list[str]] = []
	for row in raw_table:
		cleaned_row = [_clean_cell(cell) for cell in row]
		if any(cleaned_row):
			cleaned_rows.append(cleaned_row)

	if not cleaned_rows:
		return pd.DataFrame()

	max_columns = max(len(row) for row in cleaned_rows)
	normalized_rows = [row + [""] * (max_columns - len(row)) for row in cleaned_rows]

	if len(normalized_rows) == 1:
		columns = [f"coluna_{i + 1}" for i in range(max_columns)]
		return pd.DataFrame(normalized_rows, columns=columns)

	headers = _normalize_headers(normalized_rows[0], max_columns)
	return pd.DataFrame(normalized_rows[1:], columns=headers)


def _sanitize_sheet_name(name: str, used_names: set[str]) -> str:
	invalid_chars = set('[]:*?/\\')
	clean_name = "".join("_" if ch in invalid_chars else ch for ch in name).strip()
	clean_name = clean_name or "Planilha"
	clean_name = clean_name[:31]

	if clean_name not in used_names:
		used_names.add(clean_name)
		return clean_name

	base = clean_name[:28]
	index = 2
	while True:
		candidate = f"{base}_{index}"[:31]
		if candidate not in used_names:
			used_names.add(candidate)
			return candidate
		index += 1


def _normalize_municipio(raw_name: str) -> str:
	return normalize_string_for_filename(raw_name)


def _extract_info_from_new_qip_text(text: str) -> tuple[str, str, str] | None:
	match_mes_ano = _RE_MES_ANO_REFERENCIA.search(text)
	match_local = _RE_LOCAL.search(text)
	_RE_ANTIGO_HEADER = re.compile(
		r"^([A-Za-zçÇãõéêíóúâôáàéèíìóòúù\s]+)\s+([0-9]{2})\s+([0-9]{4})$"
	)

	if not match_mes_ano or not match_local:
		return None

	mes_extenso = match_mes_ano.group(1).strip().lower()
	mes = _MESES.get(mes_extenso)
	ano = match_mes_ano.group(2)
	municipio = normalize_string_for_filename(match_local.group(1).strip())

	if mes and ano and municipio:
		return mes, ano, municipio
	return None


def _extract_old_qip_payload(first_table: list[list[object]]) -> tuple[str, str, str, list[str], pd.DataFrame] | None:
	if not first_table or not first_table[0]:
		return None

	first_cell = str(first_table[0][0]).strip() if first_table[0][0] else ""
	match = re.compile(r"^([A-Za-zçÇãõéêíóúâôáàéèíìóòúù\s]+)\s+([0-9]{2})\s+([0-9]{4})$").match(first_cell)
	if not match:
		return None

	municipio = _normalize_municipio(match.group(1).strip())
	mes = match.group(2).zfill(2)
	ano = match.group(3)

	if len(first_table) < 2:
		return None

	headers = [clean_cell(cell) for cell in first_table[1]]
	data_rows: list[list[str]] = []
	for row in first_table[2:]:
		cleaned_row = [_clean_cell(cell) for cell in row]
		if any(cleaned_row):
			data_rows.append(cleaned_row)

	if not headers or not data_rows:
		return None

	return mes, ano, municipio, headers, pd.DataFrame(data_rows, columns=headers)


def _is_old_qip_header_row(row: list[str], headers: list[str]) -> bool:
	if not row or not headers:
		return False

	normalized_row = [cell.strip().lower() for cell in row]
	normalized_headers = [cell.strip().lower() for cell in headers]

	if len(normalized_row) < len(normalized_headers):
		normalized_row += [""] * (len(normalized_headers) - len(normalized_row))
	elif len(normalized_row) > len(normalized_headers):
		normalized_row = normalized_row[: len(normalized_headers)]

	return normalized_row == normalized_headers


def _old_table_to_dataframe(raw_table: list[list[object]], headers: list[str]) -> pd.DataFrame:
	rows: list[list[str]] = []
	for row in raw_table:
		cleaned_row = [clean_cell(cell) for cell in row]
		if any(cleaned_row):
			rows.append(cleaned_row)

	if not rows:
		return pd.DataFrame(columns=headers)

	first_cell = rows[0][0] if rows[0] else ""
	if first_cell and re.compile(r"^([A-Za-zçÇãõéêíóúâôáàéèíìóòúù\s]+)\s+([0-9]{2})\s+([0-9]{4})$").match(first_cell):
		rows = rows[1:]

	if rows and _is_old_qip_header_row(rows[0], headers):
		rows = rows[1:]

	if not rows:
		return pd.DataFrame(columns=headers)

	normalized_rows: list[list[str]] = []
	for row in rows:
		if len(row) < len(headers):
			normalized_rows.append(row + [""] * (len(headers) - len(row)))
		else:
			normalized_rows.append(row[: len(headers)])

	return pd.DataFrame(normalized_rows, columns=headers)


def _extract_old_qip_multpage_payload(pdf) -> tuple[str, str, str, pd.DataFrame] | None:
	first_page_tables = pdf.pages[0].extract_tables() or []
	if not first_page_tables:
		return None

	base_payload = _extract_old_qip_payload(first_page_tables[0])
	if base_payload is None:
		return None

	mes, ano, municipio, headers, first_dataframe = base_payload
	all_tables: list[pd.DataFrame] = [first_dataframe]

	for page_index, page in enumerate(pdf.pages):
		page_tables = page.extract_tables() or []
		for table_index, raw_table in enumerate(page_tables):
			if page_index == 0 and table_index == 0:
				continue

			dataframe = _old_table_to_dataframe(raw_table, headers)
			if not dataframe.empty:
				all_tables.append(dataframe)

	merged = _merge_tables_for_pdf(all_tables)
	if merged.empty:
		return None

	return mes, ano, municipio, merged


def _extract_pdf_payload(pdf_path: Path) -> tuple[str, str, str, pd.DataFrame] | None:
	try:
		with pdfplumber.open(pdf_path) as pdf:
			if not pdf.pages:
				return None

			old_payload = _extract_old_qip_multpage_payload(pdf)
			if old_payload is not None:
				return old_payload

			first_page = pdf.pages[0]

			text = first_page.extract_text() or ""
			new_info = _extract_info_from_new_qip_text(text)
			if not new_info:
				return None

			mes, ano, municipio = new_info
			tables: list[pd.DataFrame] = []
			for page in pdf.pages:
				page_tables = page.extract_tables() or []
				for raw_table in page_tables:
					if not raw_table:
						continue

					dataframe = _table_to_dataframe(raw_table)
					if not dataframe.empty:
						tables.append(dataframe)

			if not tables:
				return None

			return mes, ano, municipio, _merge_tables_for_pdf(tables)
	except Exception:
		return None


def _merge_tables_for_pdf(tables: list[pd.DataFrame]) -> pd.DataFrame:
	if not tables:
		return pd.DataFrame()

	max_columns = max(len(table.columns) for table in tables)
	base_headers = list(tables[0].columns)
	if len(base_headers) < max_columns:
		for index in range(len(base_headers), max_columns):
			base_headers.append(f"coluna_{index + 1}")

	merged_tables: list[pd.DataFrame] = []
	for table in tables:
		rows = table.astype(str).fillna("").values.tolist()
		normalized_rows = [row + [""] * (max_columns - len(row)) for row in rows]
		normalized_table = pd.DataFrame(normalized_rows, columns=base_headers)
		merged_tables.append(normalized_table)

	if not merged_tables:
		return pd.DataFrame(columns=base_headers)

	merged = pd.concat(merged_tables, ignore_index=True)
	return merged


def _apply_worksheet_style(worksheet) -> None:
	from openpyxl.utils import get_column_letter

	header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
	header_font = Font(color="FFFFFF", bold=True)
	data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
	thin_border = Border(
		left=Side(style="thin", color="D9D9D9"),
		right=Side(style="thin", color="D9D9D9"),
		top=Side(style="thin", color="D9D9D9"),
		bottom=Side(style="thin", color="D9D9D9"),
	)

	max_row = worksheet.max_row
	max_col = worksheet.max_column
	if max_row < 1 or max_col < 1:
		return

	# Detecta tipo de aba: Antigo (primeira célula da primeira linha é tipo texto com mês/ano) ou Novo (cabeçalho padrão)
	first_row = [worksheet.cell(row=1, column=col).value for col in range(1, max_col+1)]
	aba_tipo = "novo"
	if first_row[0]:
		if re.match(r"^[A-Z_]+_\d{2}_\d{4}$", worksheet.title):
			# nome da aba já padronizado
			pass
		# Heurística: se primeira célula da primeira linha é tipo da lâmpada, é antigo
		if str(first_row[0]).strip().lower() in ["tipo da lâmpada", "tipo da lampada"]:
			aba_tipo = "antigo"
		# Se segunda coluna é "Código Lâmpada" ou similar, também é antigo
		if max_col > 1 and str(first_row[1]).strip().lower() in ["código lâmpada", "codigo lampada"]:
			aba_tipo = "antigo"

	# 6. Para QIP novo: encontra a linha "Código da lâmpada"
	# e desloca tudo que está acima dela 8 colunas para a direita.
	if aba_tipo == "novo" and max_row >= 1 and max_col >= 1:
		header_row = None
		for row_index in range(1, max_row + 1):
			for col_index in range(1, max_col + 1):
				cell_value = worksheet.cell(row=row_index, column=col_index).value
				if cell_value is None:
					continue

				normalized = str(cell_value).strip().lower().replace(" ", "")
				if normalized in {"códigodalâmpada", "codigodalampada", "códigolâmpada", "codigolampada"}:
					header_row = row_index
					break
			if header_row is not None:
				break

		if header_row is not None and header_row > 1:
			current_max_col = worksheet.max_column
			for row_index in range(1, header_row):
				for col_index in range(current_max_col, 0, -1):
					origin_cell = worksheet.cell(row=row_index, column=col_index)
					destination_cell = worksheet.cell(row=row_index, column=col_index + 8)
					destination_cell.value = origin_cell.value
					origin_cell.value = None

			# Move a partir da linha do cabeçalho para cima, limitado a A:H,
			# até que a linha "Código da lâmpada" vire a linha 1.
			last_col = min(8, worksheet.max_column)
			shift_up = header_row - 1
			for row_index in range(header_row, max_row + 1):
				for col_index in range(1, last_col + 1):
					worksheet.cell(row=row_index - shift_up, column=col_index).value = worksheet.cell(
						row=row_index,
						column=col_index,
					).value

			# Limpa as últimas linhas que ficaram duplicadas após o deslocamento para cima (somente A:H).
			for row_index in range(max_row - shift_up + 1, max_row + 1):
				for col_index in range(1, last_col + 1):
					worksheet.cell(row=row_index, column=col_index).value = None

	# Recalcula limites após possíveis alterações de layout
	max_row = worksheet.max_row
	max_col = worksheet.max_column

	def _normalize_header_text(value: object) -> str:
		if value is None:
			return ""
		return str(value).strip().lower().replace(" ", "")

	# Padroniza a ordem das colunas para sempre manter
	# "Código da lâmpada" antes de "Tipo da lâmpada".
	code_headers = {"códigodalâmpada", "codigodalampada", "códigolâmpada", "codigolampada", "códigolâmpada", "codigolampada"}
	type_headers = {"tipodalâmpada", "tipodalampada", "tipodalâmpada", "tipodalampada"}

	code_col = None
	type_col = None
	for col_index in range(1, max_col + 1):
		header_value = _normalize_header_text(worksheet.cell(row=1, column=col_index).value)
		if header_value in code_headers and code_col is None:
			code_col = col_index
		if header_value in type_headers and type_col is None:
			type_col = col_index

	if code_col is not None and type_col is not None and code_col > type_col:
		for row_index in range(1, max_row + 1):
			left_cell = worksheet.cell(row=row_index, column=type_col)
			right_cell = worksheet.cell(row=row_index, column=code_col)
			left_cell.value, right_cell.value = right_cell.value, left_cell.value

	worksheet.freeze_panes = "A2"
	worksheet.auto_filter.ref = worksheet.dimensions

	# 1. Converter células numéricas para tipo numérico real
	for row_index in range(2, max_row + 1):
		for col_index in range(1, max_col + 1):
			cell = worksheet.cell(row=row_index, column=col_index)
			val = cell.value
			if isinstance(val, str):
				val_limpo = val.strip()
				if not val_limpo or not _NUM_REGEX.match(val_limpo):
					continue

				val_sem_espaco = val_limpo.replace(" ", "")
				if aba_tipo == "antigo":
					# Regra solicitada para antigo: remove vírgulas e troca ponto por vírgula
					val_corrigido = val_sem_espaco.replace(",", "").replace(".", ",")
					normalizado = val_corrigido.replace(",", ".")
				else:
					# Para novo, só normaliza separador sem mexer em textos
					if "," in val_sem_espaco and "." in val_sem_espaco:
						normalizado = val_sem_espaco.replace(".", "").replace(",", ".")
					elif "," in val_sem_espaco:
						normalizado = val_sem_espaco.replace(",", ".")
					else:
						normalizado = val_sem_espaco

				try:
					if "." in normalizado:
						cell.value = float(normalizado)
					else:
						cell.value = int(normalizado)
				except Exception:
					pass

	# 2. Centralizar horizontalmente células numéricas, 3. Centralizar verticalmente todas
	for row_index in range(1, max_row + 1):
		for col_index in range(1, max_col + 1):
			cell = worksheet.cell(row=row_index, column=col_index)
			val = cell.value
			if isinstance(val, (int, float)):
				cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			else:
				cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
			cell.border = thin_border
			if row_index == 1:
				cell.fill = header_fill
				cell.font = header_font
			elif row_index % 2 == 0:
				cell.fill = data_fill

	# 4. Primeira linha altura 30
	worksheet.row_dimensions[1].height = 30

	# 5. Largura das colunas
	tipo_coluna = None
	for col_index in range(1, max_col + 1):
		header_value = _normalize_header_text(worksheet.cell(row=1, column=col_index).value)
		if header_value in type_headers:
			tipo_coluna = col_index
			break

	for col_index in range(1, max_col + 1):
		col_letter = get_column_letter(col_index)
		worksheet.column_dimensions[col_letter].width = 30 if col_index == tipo_coluna else 15


def _apply_resumo_style(worksheet) -> None:
	from openpyxl.utils import get_column_letter

	header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
	header_font = Font(color="FFFFFF", bold=True)
	thin_border = Border(
		left=Side(style="thin", color="D9D9D9"),
		right=Side(style="thin", color="D9D9D9"),
		top=Side(style="thin", color="D9D9D9"),
		bottom=Side(style="thin", color="D9D9D9"),
	)
	status_ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
	status_ausente_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

	max_row = worksheet.max_row
	max_col = worksheet.max_column
	if max_row < 1 or max_col < 1:
		return

	worksheet.freeze_panes = "A2"
	worksheet.auto_filter.ref = worksheet.dimensions

	status_col = None
	for col_index in range(1, max_col + 1):
		header_value = worksheet.cell(row=1, column=col_index).value
		if str(header_value).strip().lower() == "status":
			status_col = col_index
			break

	for row_index in range(1, max_row + 1):
		for col_index in range(1, max_col + 1):
			cell = worksheet.cell(row=row_index, column=col_index)
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.border = thin_border

			if row_index == 1:
				cell.fill = header_fill
				cell.font = header_font

	if status_col is not None:
		for row_index in range(2, max_row + 1):
			status_cell = worksheet.cell(row=row_index, column=status_col)
			status_value = str(status_cell.value).strip().upper()
			if status_value == "OK":
				status_cell.fill = status_ok_fill
			elif status_value == "AUSENTE":
				status_cell.fill = status_ausente_fill

	worksheet.row_dimensions[1].height = 30

	for col_index in range(1, max_col + 1):
		col_letter = get_column_letter(col_index)
		worksheet.column_dimensions[col_letter].width = 10


def extract_tables_from_pdf(pdf_path: Path) -> list[pd.DataFrame]:
	tables: list[pd.DataFrame] = []

	with pdfplumber.open(pdf_path) as pdf:
		for page in pdf.pages:
			page_tables = page.extract_tables() or []
			for raw_table in page_tables:
				if not raw_table:
					continue

				dataframe = _table_to_dataframe(raw_table)
				if not dataframe.empty:
					tables.append(dataframe)

	return tables


def save_tables_to_excel(tables: list[pd.DataFrame], output_file: Path) -> None:
	output_file.parent.mkdir(parents=True, exist_ok=True)

	with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
		merged_table = _merge_tables_for_pdf(tables)
		sheet_name = "QIP"
		merged_table.to_excel(writer, sheet_name=sheet_name, index=False)
		_apply_worksheet_style(writer.sheets[sheet_name])


def save_pdfs_to_single_workbook(pdf_dataframes: dict[str, pd.DataFrame], output_file: Path) -> None:
	output_file.parent.mkdir(parents=True, exist_ok=True)
	used_sheet_names: set[str] = set()

	with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
		resumo = _build_resumo_mensal_dataframe(list(pdf_dataframes.keys()))
		resumo_sheet_name = _sanitize_sheet_name("Resumo", used_sheet_names)
		resumo.to_excel(writer, sheet_name=resumo_sheet_name, index=False)
		_apply_resumo_style(writer.sheets[resumo_sheet_name])

		for pdf_name, dataframe in pdf_dataframes.items():
			sheet_name = _sanitize_sheet_name(pdf_name, used_sheet_names)
			dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
			_apply_worksheet_style(writer.sheets[sheet_name])


def _build_resumo_mensal_dataframe(sheet_names: list[str], start_year: int = 2015) -> pd.DataFrame:
	qip_disponiveis: set[tuple[int, int]] = set()
	for sheet_name in sheet_names:
		match = re.match(r"^(\d{4})_(\d{2})_", str(sheet_name))
		if not match:
			continue

		ano = int(match.group(1))
		mes = int(match.group(2))
		if 1 <= mes <= 12:
			qip_disponiveis.add((ano, mes))

	hoje = date.today()
	linhas: list[dict[str, object]] = []

	for ano in range(hoje.year, start_year - 1, -1):
		mes_inicial = hoje.month if ano == hoje.year else 12
		for mes in range(mes_inicial, 0, -1):
			tem_qip = (ano, mes) in qip_disponiveis # type: ignore
			linhas.append({
				"Ano": ano,
				"Mes": NUM_TO_MONTH_ABBR[mes], # type: ignore
				"Mes_numero": f"{mes:02d}",
				"Status": "OK" if tem_qip else "AUSENTE",
			})

	return pd.DataFrame(linhas, columns=["Ano", "Mes", "Mes_numero", "Status"])


def save_workbooks_by_municipio(
	municipio_dataframes: dict[str, dict[str, pd.DataFrame]],
	output_dir: Path,
) -> list[Path]:
	output_dir.mkdir(parents=True, exist_ok=True)
	generated_files: list[Path] = []

	for municipio, pdf_dataframes in sorted(municipio_dataframes.items()):
		output_file = output_dir / f"QIP_{municipio}.xlsx"
		save_pdfs_to_single_workbook(pdf_dataframes, output_file)
		generated_files.append(output_file)

	return generated_files


def convert_pdf_to_xlsx(pdf_file: Path, output_dir: Path) -> Path:
	tables = extract_tables_from_pdf(pdf_file)
	if not tables:
		raise ValueError(
			f"Nenhuma tabela foi encontrada no arquivo PDF: {pdf_file.name}"
		)

	output_file = output_dir / f"{pdf_file.stem}.xlsx"
	save_tables_to_excel(tables, output_file)
	return output_file


def _rename_pdf_if_needed(pdf_file: Path, municipio: str, ano: str, mes: str) -> Path:
	novo_nome = f"QIP-{municipio}-{ano}_{mes}.pdf"
	novo_pdf_path = pdf_file.parent / novo_nome
	if pdf_file.name == novo_nome:
		return pdf_file

	if not novo_pdf_path.exists():
		pdf_file.rename(novo_pdf_path)
		return novo_pdf_path

	return pdf_file


def process_folder(input_folder: Path, output_dir: Path) -> list[Path]:
	pdf_files = sorted(input_folder.glob("*.pdf"))
	if not pdf_files:
		raise FileNotFoundError(
			f"Nenhum PDF encontrado na pasta informada: {input_folder}"
		)

	municipio_dataframes: dict[str, dict[str, pd.DataFrame]] = {}
	total_pdfs = len(pdf_files)
	print(f"Iniciando processamento de {total_pdfs} PDF(s)...")

	def process_single_pdf(pdf_file):
		try:
			payload = _extract_pdf_payload(pdf_file)
			if not payload:
				return (None, pdf_file, None, f"[AVISO] Nao foi possivel extrair dados/tabela de '{pdf_file.name}', pulando...")

			mes, ano, municipio, dataframe = payload
			_rename_pdf_if_needed(pdf_file, municipio, ano, mes)
			aba_nome = f"{ano}_{mes}_{municipio}"
			return (municipio, aba_nome, dataframe, None)
		except Exception as e:
			return (None, pdf_file, None, f"[ERRO] Falha inesperada em {pdf_file.name}: {e}")

	with ThreadPoolExecutor(max_workers=_MAX_WORKERS) as executor:
		future_to_pdf = {executor.submit(process_single_pdf, pdf_file): pdf_file for pdf_file in pdf_files}
		for idx, future in enumerate(as_completed(future_to_pdf), 1):
			municipio, aba_nome, dataframe, error = future.result()
			render_progress(idx, total_pdfs, prefix="Processando PDFs")
			if error:
				print(error)
				continue
			
			if municipio not in municipio_dataframes:
				municipio_dataframes[municipio] = {}

			used_sheet_names = set(municipio_dataframes[municipio].keys())
			aba_nome = _sanitize_sheet_name(str(aba_nome), used_sheet_names)
			municipio_dataframes[municipio][aba_nome] = dataframe
			_render_progress(idx, total_pdfs, prefix="Processando PDFs")
	if not municipio_dataframes:
		raise ValueError("Nenhuma tabela válida foi extraída dos PDFs encontrados.")

	return save_workbooks_by_municipio(municipio_dataframes, output_dir)


def build_parser() -> argparse.ArgumentParser:
	parser = argparse.ArgumentParser(
		description="Converte tabelas de PDFs de QIP para arquivos XLSX."
	)
	parser.add_argument(
		"--pdf",
		type=Path,
		help="Caminho de um PDF específico para converter.",
	)
	parser.add_argument(
		"--pasta",
		"--testes",
		type=Path,
		default=DEFAULT_INPUT_DIR,
		help="Pasta onde estão os PDFs para processar (padrão: pasta do QIPer.py).",
	)
	parser.add_argument(
		"--saida",
		type=Path,
		default=OUTPUT_DIR,
		help="Pasta de saída dos arquivos XLSX.",
	)
	parser.add_argument(
		"--arquivo-saida",
		type=str,
		default=DEFAULT_WORKBOOK_NAME,
		help="Nome do arquivo XLSX gerado quando processar pasta de testes.",
	)
	return parser


def main() -> None:
	parser = build_parser()
	args = parser.parse_args()

	if args.pdf:
		pdf_file = args.pdf.resolve()
		output_file = convert_pdf_to_xlsx(pdf_file, args.saida.resolve())
		print(f"Arquivo gerado: {output_file}")
		return

	output_dir = args.saida.resolve()
	generated_files = process_folder(args.pasta.resolve(), output_dir)

	if args.arquivo_saida != DEFAULT_WORKBOOK_NAME and len(generated_files) == 1:
		custom_output = output_dir / args.arquivo_saida
		generated_files[0].replace(custom_output)
		generated_files = [custom_output]

	print("Arquivos gerados por município:")
	for file_path in generated_files:
		print(f"- {file_path}")


if __name__ == "__main__":
	main()
