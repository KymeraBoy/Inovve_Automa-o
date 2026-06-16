from __future__ import annotations

import argparse
from datetime import datetime
import re
import unicodedata
from dataclasses import dataclass

import fitz
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter # type: ignore
from pathlib import Path

from utils import MONTH_MAPPINGS, MONTH_ABBR_TO_NUM, NUM_TO_MONTH_ABBR, normalize_string_for_filename, strip_accents, INVALID_FILENAME_CHARS, build_unique_path


@dataclass
class InvoiceData:
	municipio: str
	mes_ano: str
	unidade_consumidora: str


def month_year_sort_key(value: str) -> tuple[int, int]:
	match = re.fullmatch(r"([A-Z]{3})_(\d{4})", value) # type: ignore
	if not match:
		return (9999, 99)
	month_abbr, year = match.groups()
	month_index = MONTH_ABBR_TO_NUM.get(month_abbr, 99)
	return (int(year), month_index) # type: ignore


def read_pdf_text(pdf_path: Path) -> str:
	with fitz.open(pdf_path) as doc:
		return "\n".join(page.get_text() for page in doc)


def extract_municipio(text: str) -> str | None:
	lines = [line.strip() for line in text.splitlines() if line.strip()]

	for line in lines:
		match = re.search(r"PREFEITURA\s+MUNICIPAL\s+DE\s+(.+)$", line, re.IGNORECASE)
		if not match:
			continue
		municipio = match.group(1).strip() # type: ignore
		municipio = re.sub(r"\s+PB$", "", municipio, flags=re.IGNORECASE)
		municipio = re.sub(r"\s+\([A-Z]{2}:.*$", "", municipio)
		municipio = re.sub(r"\s+", " ", municipio)
		municipio = normalize_for_filename(municipio)
		if municipio:
			return municipio

	def infer_from_delivery_line(line: str) -> str | None:
		norm = strip_accents(line).upper()
		norm = re.sub(r"\s+", " ", norm).strip()

		# Ex.: "COREMAS (AG: 227)"
		ag_match = re.search(r"\b([A-Z]{3,})\s*\(AG:\s*\d+\)", norm)
		if ag_match:
			candidate = normalize_string_for_filename(ag_match.group(1)) # type: ignore
			if candidate:
				return candidate

		# Ex.: "PM COREMAS ..." / "PM DE COREMAS ..." / "P.M. COREMAS ..."
		pm_match = re.search(r"\bP\.?\s*M\.?\s*(?:DE\s+|DO\s+|DA\s+)?([A-Z]{3,})\b", norm)
		if pm_match:
			candidate = normalize_string_for_filename(pm_match.group(1)) # type: ignore
			if candidate:
				return candidate

		return None

	# Fallback para layout onde o destino vem após "DOMICILIO DE ENTREGA".
	for idx, line in enumerate(lines):
		norm = strip_accents(line).upper()
		if "DOMIC" in norm and "ENTREGA" in norm:
			for offset in (1, 2, 3):
				if idx + offset >= len(lines):
					break
				candidate = infer_from_delivery_line(lines[idx + offset])
				if candidate:
					return candidate

	# Fallback adicional: busca direta por padrões PM no documento inteiro.
	for line in lines:
		candidate = infer_from_delivery_line(line)
		if candidate:
			return candidate

	return None


def extract_mes_ano(text: str) -> str | None:
	full_name_pattern = re.compile(
		r"(janeiro|fevereiro|marco|março|abril|maio|junho|julho|agosto|setembro|outubro|novembro|dezembro)\s*/\s*(\d{4})",
		re.IGNORECASE,
	)
	match = full_name_pattern.search(text)
	if match: # type: ignore
		mes = MONTH_MAPPINGS.get(match.group(1).lower(), {}).get("abbr") # type: ignore
		ano = match.group(2) # type: ignore
		if mes:
			return f"{mes}_{ano}"

	abbr_pattern = re.compile(r"\b(JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)\s*/\s*(\d{2})\b", re.IGNORECASE)
	match = abbr_pattern.search(text)
	if match: # type: ignore
		mes = match.group(1).upper() # type: ignore
		ano_2 = match.group(2) # type: ignore
		if mes in MONTH_ABBR_TO_NUM: # type: ignore
			return f"{mes}_20{ano_2}"

	return None


def normalize_uc(raw_uc: str) -> str:
	def strip_leading_zeros(number: str) -> str:
		cleaned = number.lstrip("0")
		return cleaned or "0"

	raw_uc = raw_uc.strip()
	if "/" in raw_uc:
		raw_uc = raw_uc.split("/", maxsplit=1)[1]
	raw_uc = raw_uc.replace(" ", "")

	match_with_dv = re.fullmatch(r"(\d+)-(\d)", raw_uc)
	if match_with_dv:
		numero, digito = match_with_dv.groups()
		numero = strip_leading_zeros(numero)
		return f"{numero}-{digito}"

	match_without_dv = re.fullmatch(r"\d{8,10}", raw_uc)
	if match_without_dv:
		# Alguns PDFs chegam sem o dígito final; removemos zeros à esquerda.
		return strip_leading_zeros(raw_uc)

	return raw_uc


def extract_unidade_consumidora(text: str) -> str | None:
	# Nas faturas novas, a UC costuma aparecer como "5/25806-1" no corpo do texto.
	# Priorizar esse padrão evita capturar a matrícula longa (ex.: 25806-2025-9-7).
	priority_global_patterns = [
		re.compile(r"\b(\d+/\d{4,}-\d)\b"),
		re.compile(r"\b(\d{10}-\d)\b"),
	]
	for pattern in priority_global_patterns:
		match = pattern.search(text)
		if match:
			return normalize_uc(match.group(1))

	patterns_near_matricula = [
		re.compile(r"\b(\d{10}-\d)\b"),
		re.compile(r"\b(\d+/\d{4,}-\d)\b"),
		re.compile(r"\b(\d{10})\b"),
		re.compile(r"\b(\d+/\d{4,10})\b"),
		re.compile(r"\b(\d{4,6}-\d{4}-\d-\d)\b"),
	]
	patterns_global = [
		re.compile(r"\b(\d{10}-\d)\b"),
		re.compile(r"\b(\d+/\d{4,}-\d)\b"),
		re.compile(r"\b(\d{4,6}-\d{4}-\d-\d)\b"),
	]

	lines = [line.strip() for line in text.splitlines() if line.strip()]
	matricula_indexes = [idx for idx, line in enumerate(lines) if "MATR" in strip_accents(line).upper()]

	# Procura primeiro perto de "MATRICULA", onde normalmente a UC aparece.
	for index in matricula_indexes:
		window = lines[index : index + 10]
		joined = "\n".join(window)
		for pattern in patterns_near_matricula:
			match = pattern.search(joined)
			if match:
				return normalize_uc(match.group(1))

	for pattern in patterns_global:
		match = pattern.search(text)
		if match:
			return normalize_uc(match.group(1))

	return None


def extract_invoice_data(pdf_path: Path) -> InvoiceData | None:
	text = read_pdf_text(pdf_path)

	municipio = extract_municipio(text)
	mes_ano = extract_mes_ano(text)
	unidade = extract_unidade_consumidora(text)

	if not (municipio and mes_ano and unidade):
		return None

	return InvoiceData(municipio=municipio, mes_ano=mes_ano, unidade_consumidora=unidade)


def build_month_year_columns(found_months: set[str]) -> list[str]:
	valid_months = [m for m in found_months if re.fullmatch(r"[A-Z]{3}_\d{4}", m)]
	if not valid_months:
		return []

	years = [int(m.split("_")[1]) for m in valid_months]
	first_year = min(years) # type: ignore
	last_year = max(years)

	columns: list[str] = []
	for year in range(first_year, last_year + 1):
		for month in MONTH_ORDER:
			columns.append(f"{month}_{year}")
	return columns
MONTH_ORDER = [NUM_TO_MONTH_ABBR[i] for i in range(1, 13)] # type: ignore

def generate_status_spreadsheet(records: list[InvoiceData], output_folder: Path) -> Path | None:
	if not records:
		return None

	all_months = {item.mes_ano for item in records}
	month_columns = build_month_year_columns(all_months)

	status_map: dict[tuple[str, str], set[str]] = {}
	for item in records:
		key = (item.municipio, item.unidade_consumidora)
		status_map.setdefault(key, set()).add(item.mes_ano)

	rows: list[dict[str, str | int]] = []
	for (municipio, uc), present_months in sorted(status_map.items(), key=lambda k: (k[0][0], k[0][1])):
		row: dict[str, str | int] = {
			"MUNICIPIO": municipio,
			"UNIDADE_CONSUMIDORA": uc,
		}

		missing_months: list[str] = []
		for month_col in month_columns:
			if month_col in present_months:
				row[month_col] = "OK"
			else:
				row[month_col] = ""
				missing_months.append(month_col)

		row["QTD_FALTANTES"] = len(missing_months)
		row["FALTANTES"] = ", ".join(missing_months)
		rows.append(row)

	ordered_columns = ["MUNICIPIO", "UNIDADE_CONSUMIDORA", *month_columns, "QTD_FALTANTES", "FALTANTES"]
	df = pd.DataFrame(rows, columns=ordered_columns)

	xlsx_path = output_folder / "Status_Faturas.xlsx"
	csv_path = output_folder / "Status_Faturas.csv"
	try:
		df.to_excel(xlsx_path, index=False)
		df.to_csv(csv_path, index=False, encoding="utf-8-sig")
		format_status_workbook(xlsx_path, month_columns)
	except PermissionError:
		timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
		xlsx_path = output_folder / f"Status_Faturas_{timestamp}.xlsx"
		csv_path = output_folder / f"Status_Faturas_{timestamp}.csv"
		df.to_excel(xlsx_path, index=False)
		df.to_csv(csv_path, index=False, encoding="utf-8-sig")
		format_status_workbook(xlsx_path, month_columns)

	return xlsx_path


def format_status_workbook(xlsx_path: Path, month_columns: list[str]) -> None:
	wb = load_workbook(xlsx_path)
	ws = wb.active
	ws.title = "STATUS"

	header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
	header_font = Font(color="FFFFFF", bold=True)
	header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

	ok_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
	missing_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")

	max_row = ws.max_row
	max_col = ws.max_column

	for col in range(1, max_col + 1):
		cell = ws.cell(row=1, column=col)
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = header_alignment

	ws.freeze_panes = "C2"
	ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

	month_start_col = 3
	month_end_col = month_start_col + max(len(month_columns) - 1, 0)

	ws.column_dimensions["A"].width = 18
	ws.column_dimensions["B"].width = 22
	for col in range(month_start_col, month_end_col + 1):
		ws.column_dimensions[get_column_letter(col)].width = 10

	qtd_col = month_end_col + 1
	faltantes_col = month_end_col + 2
	ws.column_dimensions[get_column_letter(qtd_col)].width = 14
	ws.column_dimensions[get_column_letter(faltantes_col)].width = 56

	for row in range(2, max_row + 1):
		for col in range(1, max_col + 1):
			cell = ws.cell(row=row, column=col)
			if col in (1, 2, faltantes_col):
				cell.alignment = Alignment(horizontal="left", vertical="center")
			else:
				cell.alignment = Alignment(horizontal="center", vertical="center")

	if max_row >= 2 and month_end_col >= month_start_col:
		month_range = f"{get_column_letter(month_start_col)}2:{get_column_letter(month_end_col)}{max_row}"
		ws.conditional_formatting.add(
			month_range,
			CellIsRule(operator="equal", formula=['"OK"'], fill=ok_fill, stopIfTrue=True),
		)
		ws.conditional_formatting.add(
			month_range,
			CellIsRule(operator="notEqual", formula=['"OK"'], fill=missing_fill),
		)

	qtd_range = f"{get_column_letter(qtd_col)}2:{get_column_letter(qtd_col)}{max_row}"
	high_missing_rule = CellIsRule(operator="greaterThanOrEqual", formula=["3"], fill=PatternFill(fill_type="solid", fgColor="F8CBAD"))
	ws.conditional_formatting.add(qtd_range, high_missing_rule)

	wb.save(xlsx_path)


def process_folder(folder: Path, dry_run: bool) -> None:
	pdf_files = sorted(
		(p for p in folder.rglob("*") if p.is_file() and p.suffix.lower() == ".pdf"),
		key=lambda p: str(p).lower(),
	)

	if not pdf_files:
		print("Nenhum PDF encontrado na pasta.")
		return

	success = 0
	failures: list[Path] = []
	processed_records: list[InvoiceData] = []

	for pdf_path in pdf_files:
		data = extract_invoice_data(pdf_path)
		if data is None:
			failures.append(pdf_path)
			print(f"[FALHA] Nao foi possivel extrair dados de: {pdf_path.name}")
			continue

		processed_records.append(data)

		base_name = f"{data.municipio}-{data.mes_ano}-{data.unidade_consumidora}.pdf"
		base_name = INVALID_FILENAME_CHARS.sub("-", base_name)
		desired_path = pdf_path.with_name(base_name)

		if desired_path == pdf_path:
			print(f"[OK] Ja esta no formato: {pdf_path.name}")
			success += 1
			continue

		target_path = build_unique_path(desired_path)

		if dry_run:
			print(f"[SIMULACAO] {pdf_path.name} -> {target_path.name}")
		else:
			pdf_path.rename(target_path)
			print(f"[RENOMEADO] {pdf_path.name} -> {target_path.name}")

		success += 1

	print("\nResumo")
	print(f"Total de PDFs: {len(pdf_files)}")
	print(f"Processados com sucesso: {success}")
	print(f"Falhas: {len(failures)}")

	if failures:
		print("Arquivos com falha:")
		for failed in failures:
			print(f"- {failed.name}")

	spreadsheet_path = generate_status_spreadsheet(processed_records, folder)
	if spreadsheet_path is not None:
		print(f"\nPlanilha de status gerada em: {spreadsheet_path.name}")


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(
		description=(
			"Organiza e renomeia faturas PDF da pasta atual no formato: "
			"NOME_DO_MUNICIPIO-MES_ANO-UNIDADE_CONSUMIDORA.pdf"
		)
	)
	parser.add_argument(
		"--folder",
		type=Path,
		default=Path(__file__).resolve().parent,
		help="Pasta dos PDFs. Por padrao usa a pasta onde o script esta.",
	)
	parser.add_argument(
		"--simular",
		action="store_true",
		help="Somente simula. Por padrao, o script renomeia os arquivos de fato.",
	)
	return parser.parse_args()


def main() -> None:
	args = parse_args()
	folder = args.folder.resolve()
	process_folder(folder, dry_run=args.simular)


if __name__ == "__main__":
	main()
