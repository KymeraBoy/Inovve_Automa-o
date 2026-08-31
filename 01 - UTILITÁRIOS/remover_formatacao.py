import os
import shutil
import tempfile

import win32com.client as win32
from openpyxl import load_workbook, Workbook


# =============================================================
# CONFIGURAÇÃO
# =============================================================

# Pasta onde estão os arquivos XLSX originais
PASTA_ENTRADA = r"C:\Users\Usuário 1\OneDrive\00 - Arquivos antigos\Documentos\Arthur Castro\QIP - Itaíba\Layout 02"

# Pasta onde serão salvos os arquivos limpos
PASTA_SAIDA = os.path.join(
    PASTA_ENTRADA,
    "PROCESSADOS"
)


# =============================================================
# FUNÇÕES AUXILIARES
# =============================================================

def obter_arquivos_xlsx(pasta):
    """
    Retorna todos os arquivos .xlsx encontrados diretamente
    na pasta informada.

    Ignora arquivos temporários do Excel (~$).
    """

    arquivos = []

    for nome in os.listdir(pasta):

        caminho = os.path.join(
            pasta,
            nome
        )

        # Ignora pastas
        if not os.path.isfile(caminho):
            continue

        # Ignora arquivos temporários do Excel
        if nome.startswith("~$"):
            continue

        # Aceita somente arquivos XLSX
        if not nome.lower().endswith(".xlsx"):
            continue

        arquivos.append(caminho)

    return sorted(arquivos)


# =============================================================
# RECALCULAR ARQUIVO COM O MICROSOFT EXCEL
# =============================================================

def recalcular_com_excel(arquivo_entrada):

    """
    Abre o arquivo no Microsoft Excel, força o recálculo
    completo das fórmulas e salva uma cópia temporária.

    Essa etapa é importante para garantir que fórmulas como:

        F17 = =E19
        E19 = itaiba

    sejam convertidas para:

        F17 = itaiba
    """

    pasta_temp = tempfile.mkdtemp()

    arquivo_recalculado = os.path.join(
        pasta_temp,
        "arquivo_recalculado.xlsx"
    )

    excel = None
    workbook = None

    try:

        print("  Abrindo no Microsoft Excel...")

        excel = win32.DispatchEx(
            "Excel.Application"
        )

        # Excel invisível
        excel.Visible = False

        # Não exibir caixas de diálogo
        excel.DisplayAlerts = False

        # Abre o arquivo original
        workbook = excel.Workbooks.Open(
            os.path.abspath(arquivo_entrada),
            UpdateLinks=0,
            ReadOnly=False
        )

        print("  Recalculando fórmulas...")

        # Força o recálculo completo
        excel.CalculateFullRebuild()

        # Aguarda consultas assíncronas
        try:
            excel.CalculateUntilAsyncQueriesDone()
        except:
            pass

        # Salva uma cópia temporária
        workbook.SaveAs(
            os.path.abspath(
                arquivo_recalculado
            ),
            FileFormat=51
        )

        return (
            arquivo_recalculado,
            pasta_temp
        )

    finally:

        # Fecha o arquivo
        if workbook is not None:

            try:
                workbook.Close(
                    SaveChanges=False
                )
            except:
                pass

        # Fecha o Excel
        if excel is not None:

            try:
                excel.Quit()
            except:
                pass


# =============================================================
# PROCESSA UM ARQUIVO
# =============================================================

def processar_arquivo(
    arquivo_entrada,
    pasta_saida
):

    nome_arquivo = os.path.basename(
        arquivo_entrada
    )

    nome_base = os.path.splitext(
        nome_arquivo
    )[0]

    arquivo_saida = os.path.join(
        pasta_saida,
        f"{nome_base}_limpa.xlsx"
    )

    print()
    print("=" * 70)
    print(f"PROCESSANDO: {nome_arquivo}")
    print("=" * 70)

    pasta_temp = None

    try:

        # =====================================================
        # 1. RECALCULA USANDO O EXCEL
        # =====================================================

        (
            arquivo_recalculado,
            pasta_temp
        ) = recalcular_com_excel(
            arquivo_entrada
        )

        # =====================================================
        # 2. ABRE A PLANILHA RECALCULADA
        # =====================================================

        # Versão contendo fórmulas
        wb_formulas = load_workbook(
            arquivo_recalculado,
            data_only=False
        )

        # Versão contendo os resultados das fórmulas
        wb_valores = load_workbook(
            arquivo_recalculado,
            data_only=True
        )

        # =====================================================
        # 3. CRIA UM NOVO XLSX DO ZERO
        # =====================================================

        wb_novo = Workbook()

        # Remove a planilha criada automaticamente
        wb_novo.remove(
            wb_novo.active
        )

        # =====================================================
        # 4. PROCESSA CADA ABA
        # =====================================================

        for indice, ws_formulas in enumerate(
            wb_formulas.worksheets
        ):

            ws_valores = (
                wb_valores.worksheets[indice]
            )

            nome_aba = ws_formulas.title

            print(
                f"  Processando aba: {nome_aba}"
            )

            # -------------------------------------------------
            # CRIA UMA NOVA ABA VAZIA
            # -------------------------------------------------

            ws_nova = wb_novo.create_sheet(
                title=nome_aba
            )

            # -------------------------------------------------
            # PERCORRE TODAS AS CÉLULAS
            # -------------------------------------------------

            for linha in range(
                1,
                ws_formulas.max_row + 1
            ):

                for coluna in range(
                    1,
                    ws_formulas.max_column + 1
                ):

                    # Célula original
                    celula_original = (
                        ws_formulas.cell(
                            linha,
                            coluna
                        )
                    )

                    # Célula com o resultado calculado
                    celula_valor = (
                        ws_valores.cell(
                            linha,
                            coluna
                        )
                    )

                    # =================================================
                    # PEGA O VALOR CALCULADO
                    # =================================================

                    valor = celula_valor.value

                    # =================================================
                    # GRAVA SOMENTE O VALOR
                    # =================================================

                    if valor is not None:

                        nova_celula = (
                            ws_nova.cell(
                                linha,
                                coluna
                            )
                        )

                        nova_celula.value = valor

            print(
                f"    ✓ Aba processada"
            )

        # =====================================================
        # 5. SALVA O NOVO ARQUIVO
        # =====================================================

        wb_novo.save(
            arquivo_saida
        )

        # =====================================================
        # 6. FECHA OS ARQUIVOS
        # =====================================================

        wb_formulas.close()
        wb_valores.close()
        wb_novo.close()

        print()
        print("  ✓ SUCESSO")
        print(
            f"  Arquivo criado: {arquivo_saida}"
        )

        return True, None

    except Exception as erro:

        print()
        print("  ✗ ERRO")
        print(
            f"  {erro}"
        )

        return False, str(erro)

    finally:

        # =====================================================
        # REMOVE ARQUIVOS TEMPORÁRIOS
        # =====================================================

        if pasta_temp is not None:

            shutil.rmtree(
                pasta_temp,
                ignore_errors=True
            )


# =============================================================
# PROCESSAMENTO DA PASTA
# =============================================================

def processar_pasta():

    print()
    print("=" * 70)
    print("LIMPADOR DE PLANILHAS EXCEL")
    print("=" * 70)
    print()

    # =========================================================
    # VERIFICA A PASTA
    # =========================================================

    if not os.path.exists(
        PASTA_ENTRADA
    ):

        print(
            "ERRO: a pasta de entrada não existe:"
        )

        print(
            PASTA_ENTRADA
        )

        return

    # =========================================================
    # CRIA A PASTA DE SAÍDA
    # =========================================================

    os.makedirs(
        PASTA_SAIDA,
        exist_ok=True
    )

    # =========================================================
    # LOCALIZA OS XLSX
    # =========================================================

    arquivos = obter_arquivos_xlsx(
        PASTA_ENTRADA
    )

    if not arquivos:

        print(
            "Nenhum arquivo .xlsx encontrado."
        )

        return

    print(
        f"Arquivos encontrados: {len(arquivos)}"
    )

    # =========================================================
    # CONTADORES
    # =========================================================

    sucessos = 0
    erros = 0

    arquivos_com_erro = []

    # =========================================================
    # PROCESSA OS ARQUIVOS
    # =========================================================

    for numero, arquivo in enumerate(
        arquivos,
        start=1
    ):

        print()
        print(
            f"[{numero}/{len(arquivos)}]"
        )

        sucesso, erro = processar_arquivo(
            arquivo,
            PASTA_SAIDA
        )

        if sucesso:

            sucessos += 1

        else:

            erros += 1

            arquivos_com_erro.append(
                (
                    os.path.basename(
                        arquivo
                    ),
                    erro
                )
            )

    # =========================================================
    # RELATÓRIO FINAL
    # =========================================================

    print()
    print()
    print("=" * 70)
    print("PROCESSAMENTO FINALIZADO")
    print("=" * 70)

    print()
    print(
        f"Total de arquivos: {len(arquivos)}"
    )

    print(
        f"Processados com sucesso: {sucessos}"
    )

    print(
        f"Com erro: {erros}"
    )

    # =========================================================
    # MOSTRA ERROS
    # =========================================================

    if arquivos_com_erro:

        print()
        print("ARQUIVOS COM ERRO:")
        print()

        for nome, erro in arquivos_com_erro:

            print(
                f"- {nome}"
            )

            print(
                f"  {erro}"
            )

    # =========================================================
    # LOCAL DOS RESULTADOS
    # =========================================================

    print()
    print(
        "Arquivos limpos:"
    )

    print(
        PASTA_SAIDA
    )

    print()
    print("=" * 70)


# =============================================================
# EXECUÇÃO
# =============================================================

if __name__ == "__main__":

    processar_pasta()
