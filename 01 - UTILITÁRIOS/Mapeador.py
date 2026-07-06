import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side


def mapear_estrutura_pastas(caminho_raiz):
    # Dicionário aninhado para representar a árvore de diretórios
    estrutura = {}

    # Percorre a pasta de baixo para cima (bottom-up) para facilitar a montagem
    for raiz, pastas, _ in os.walk(caminho_raiz):
        # Ignora a própria pasta raiz no mapeamento interno
        if raiz == caminho_raiz:
            partes_relativas = []
        else:
            partes_relativas = os.path.relpath(raiz, caminho_raiz).split(
                os.sep
            )

        atual = estrutura
        for parte in partes_relativas:
            if parte not in atual:
                atual[parte] = {}
            atual = atual[parte]

    return estrutura


def achatar_estrutura(dicionario_estrutura):
    # Transforma o dicionário em uma lista de caminhos (linhas)
    if not dicionario_estrutura:
        return [[]]

    linhas = []
    for pasta, subpastas in dicionario_estrutura.items():
        sub_linhas = achatar_estrutura(subpastas)
        for sub_linha in sub_linhas:
            linhas.append([pasta] + sub_linha)
    return linhas


def gerar_planilha_pastas(caminho_raiz, arquivo_saida="organizacao_pastas.xlsx"):
    if not os.path.exists(caminho_raiz):
        print(f"Erro: O caminho '{caminho_raiz}' não existe.")
        return

    print("Mapeando estrutura de pastas...")
    estrutura = mapear_estrutura_pastas(caminho_raiz)
    linhas_dados = achatar_estrutura(estrutura)

    # Se a pasta raiz estiver vazia
    if not linhas_dados or linhas_dados == [[]]:
        print("A pasta informada não contém subpastas.")
        return

    # Criando o Workbook do Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Estrutura de Pastas"

    # Estilos básicos para deixar a planilha legível
    alinhamento_central = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    borda_fina = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    cor_cabecalho = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )

    # Determina a profundidade máxima (quantidade de colunas)
    max_colunas = max(len(linha) for linha in linhas_dados)

    # Escreve o cabeçalho
    for col_idx in range(1, max_colunas + 1):
        celula = ws.cell(row=1, column=col_idx, value=f"Nível {col_idx}")
        celula.fill = cor_cabecalho
        celula.alignment = alinhamento_central
        celula.font = ws.cell(row=1, column=col_idx).font.copy(
            bold=True, color="FFFFFF"
        )

    # Escreve os dados linha por linha (iniciando na linha 2)
    for row_idx, linha in enumerate(linhas_dados, start=2):
        for col_idx, valor in enumerate(linha, start=1):
            celula = ws.cell(row=row_idx, column=col_idx, value=valor)
            celula.alignment = alinhamento_central
            celula.border = borda_fina

    # Aplica a mesclagem vertical (de cima para baixo, coluna por coluna)
    # Ignora a última coluna (nível mais baixo), conforme solicitado
    for col in range(1, max_colunas):
        row_inicio = 2
        while row_inicio <= len(linhas_dados) + 1:
            val_atual = ws.cell(row=row_inicio, column=col).value

            # Se o valor for vazio (fim da árvore local), passa para o próximo
            if val_atual is None:
                row_inicio += 1
                continue

            row_fim = row_inicio
            # Verifica até onde o mesmo valor se repete na mesma coluna
            # E garante que pertencem ao mesmo "pai" checando a coluna anterior (se houver)
            while row_fim <= len(linhas_dados) + 1:
                proximo_val = ws.cell(row=row_fim + 1, column=col).value

                # Condição para continuar a mesclagem: mesmo valor
                mesmo_valor = proximo_val == val_atual

                # Condição extra: o pai (coluna anterior) também deve ser o mesmo para evitar mesclar pastas com o mesmo nome em ramos diferentes
                mesmo_pai = True
                if col > 1:
                    pai_atual = ws.cell(row=row_inicio, column=col - 1).value
                    proximo_pai = ws.cell(row=row_fim + 1, column=col - 1).value
                    mesmo_pai = pai_atual == proximo_pai

                if mesmo_valor and mesmo_pai:
                    row_fim += 1
                else:
                    break

            # Se houver mais de uma linha com o mesmo conteúdo, mescla
            if row_fim > row_inicio:
                ws.merge_cells(
                    start_row=row_inicio,
                    start_column=col,
                    end_row=row_fim,
                    end_column=col,
                )
                # Reaplica o alinhamento na célula master após a mesclagem (o openpyxl às vezes perde)
                ws.cell(row=row_inicio, column=col).alignment = (
                    alinhamento_central
                )

            row_inicio = row_fim + 1

    # Ajusta automaticamente a largura das colunas
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    wb.save(arquivo_saida)
    print(f"Planilha gerada com sucesso: {arquivo_saida}")


# --- EXECUÇÃO DO SCRIPT ---
if __name__ == "__main__":
    # Substitua pelo caminho da pasta que você quer mapear
    # Exemplo no Windows: r"C:\Usuarios\Nome\Documentos"
    # Exemplo no Mac/Linux: "/home/usuario/documentos"
    pasta_alvo = input(
        "Digite ou cole o caminho completo da pasta: "
    ).strip('"')

    gerar_planilha_pastas(pasta_alvo)