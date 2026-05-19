import openpyxl
from pathlib import Path

xlsx_path = Path(__file__).resolve().parent / "Testes" / "Output" / "QIP_consolidado.xlsx"

print(f"Inspecionando: {xlsx_path}\n")

try:
    workbook = openpyxl.load_workbook(xlsx_path)
    print(f"Total de abas: {len(workbook.sheetnames)}\n")
    
    for i, sheet_name in enumerate(workbook.sheetnames, start=1):
        sheet = workbook[sheet_name]
        print(f"{i}. Aba: '{sheet_name}'")
        print(f"   Linhas: {sheet.max_row}, Colunas: {sheet.max_column}")
        
        # Primeiras linhas
        print(f"   Primeiras 3 linhas:")
        for row_idx in range(1, min(4, sheet.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(6, sheet.max_column + 1)):
                cell = sheet.cell(row_idx, col_idx)
                row_data.append(str(cell.value)[:20] if cell.value else "")
            print(f"     {row_data}")
        print()
except Exception as e:
    print(f"Erro ao inspecionar: {e}")
