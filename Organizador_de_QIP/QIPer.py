from __future__ import annotations

import argparse
import sys
from pathlib import Path
import re

import pandas as pd
import pdfplumber
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from concurrent.futures import ThreadPoolExecutor, as_completed


BASE_DIR = Path(__file__).resolve().parent
TESTES_DIR = BASE_DIR / "Testes"
OUTPUT_DIR = TESTES_DIR / "Output"
DEFAULT_WORKBOOK_NAME = "QIP_consolidado.xlsx"


def _render_progress(current: int, total: int, prefix: str = "Progresso") -> None:
	if total <= 0:
		return

	bar_width = 30
	ratio = current / total
	filled = int(bar_width * ratio)
	bar = "#" * filled + "-" * (bar_width - filled)
	percent = ratio * 100
	message = f"\r{prefix}: [{bar}] {current}/{total} ({percent:5.1f}%)"
	sys.stdout.write(message)
	sys.stdout.flush()

	if current >= total:
		sys.stdout.write("\n")


def _clean_cell(value: object) -> str:
	if value is None:
		return ""
	return str(value).replace("\n", " ").strip()


def _normalize_headers(headers: list[str], total_columns: int) -> list[str]:
	normalized: list[str] = []
	used_names: dict[str, int] = {}

	for index in range(total_columns):
		raw_header = headers[index] if index < len(headers) else ""
		base_name = raw_header if raw_header else f"coluna_{index + 1}"

		count = used_names.get(base_name, 0)
		used_names[base_name] = count + 1

		if count > 0:
			normalized.append(f"{base_name}_{count + 1}")
		else:
			normalized.append(base_name)

	return normalized


def _table_to_dataframe(raw_table: list[list[object]]) -> pd.DataFrame:
	cleaned_rows: list[list[str]] = []
	for row in raw_table:
		cleaned_row = [_clean_cell(cell) for cell in row]
		if any(cleaned_row):
			cleaned_rows.append(cleaned_row)

	if not cleaned_rows:
		return pd.DataFrame()

	max_columns = max(len(row) for row in cleaned_rows)
	normalized_rows = [row + [""] * (max_columns - len(row)) for row in cleaned_rows]

	if len(normalized_rows) == 1:
		columns = [f"coluna_{i + 1}" for i in range(max_columns)]
		return pd.DataFrame(normalized_rows, columns=columns)

	headers = _normalize_headers(normalized_rows[0], max_columns)
	return pd.DataFrame(normalized_rows[1:], columns=headers)


def _sanitize_sheet_name(name: str, used_names: set[str]) -> str:
	invalid_chars = set('[]:*?/\\')
	clean_name = "".join("_" if ch in invalid_chars else ch for ch in name).strip()
	clean_name = clean_name or "Planilha"
	clean_name = clean_name[:31]

	if clean_name not in used_names:
		used_names.add(clean_name)
		return clean_name

	base = clean_name[:28]
	index = 2
	while True:
		candidate = f"{base}_{index}"[:31]
		if candidate not in used_names:
			used_names.add(candidate)
			return candidate
		index += 1


# Mapeamento de meses por extenso para número
_MESES = {
	"janeiro": "01", "fevereiro": "02", "março": "03", "marco": "03", "abril": "04", "maio": "05", "junho": "06",
	"julho": "07", "agosto": "08", "setembro": "09", "outubro": "10", "novembro": "11", "dezembro": "12"
}

def _extract_mes_ano_municipio_from_pdf(pdf_path: Path) -> tuple[str, str, str] | None:
	try:
		with pdfplumber.open(pdf_path) as pdf:
			if not pdf.pages:
				return None
			text = pdf.pages[0].extract_text() or ""
	except Exception:
		return None

	mes, ano, municipio = None, None, None
	
	# Busca mês/ano: padrão "Mês/Ano de Referência Abril/2025"
	match_mes_ano = re.search(r"[Mm]ês/[Aa]no\s+de\s+[Rr]eferência\s+([A-Za-zçÇãõéêíóúâôáàéèíìóòúù]+)/(\d{4})", text)
	if match_mes_ano:
		mes_extenso = match_mes_ano.group(1).strip().lower()
		ano = match_mes_ano.group(2)
		mes = _MESES.get(mes_extenso)

	# Busca município: padrão "Local AGUIAR"
	match_local = re.search(r"Local\s+([A-Za-zçÇãõéêíóúâôáàéèíìóòúù\s]+)(?:\n|$)", text, re.IGNORECASE)
	if match_local:
		municipio = match_local.group(1).strip().upper()
		# Remove acentos e caracteres especiais, deixa só letras e espaços
		municipio = re.sub(r"[^A-Z ]", "", municipio)
		municipio = re.sub(r"\s+", "_", municipio)

	if mes and ano and municipio:
		return mes, ano, municipio
	return None


def _is_qip_antigo(pdf_path: Path) -> bool:
	try:
		with pdfplumber.open(pdf_path) as pdf:
			if not pdf.pages:
				return False
			page = pdf.pages[0]
			tables = page.extract_tables()
			if not tables:
				return False
			
			# QIP antigo: primeira célula contém padrão "MUNICIPIO MES ANO" (ex: "AGUIAR 07 2015")
			first_table = tables[0]
			if not first_table or not first_table[0]:
				return False
			
			first_cell = str(first_table[0][0]).strip() if first_table[0][0] else ""
			# Padrão: palavra espaço 2 dígitos espaço 4 dígitos
			match = re.match(r"^[A-Za-zçÇãõéêíóúâôáàéèíìóòúù\s]+\s+(\d{2})\s+(\d{4})$", first_cell)
			return match is not None
	except Exception:
		return False


def _extract_mes_ano_municipio_from_qip_antigo(pdf_path: Path) -> tuple[str, str, str] | None:
	try:
		with pdfplumber.open(pdf_path) as pdf:
			if not pdf.pages:
				return None
			page = pdf.pages[0]
			tables = page.extract_tables()
			if not tables:
				return None
			
			first_table = tables[0]
			if not first_table or not first_table[0]:
				return None
			
			# Primeira célula contém "MUNICIPIO MES ANO"
			first_cell = str(first_table[0][0]).strip() if first_table[0][0] else ""
			match = re.match(r"^([A-Za-zçÇãõéêíóúâôáàéèíìóòúù\s]+)\s+(\d{2})\s+(\d{4})$", first_cell)
			if not match:
				return None
			
			municipio = match.group(1).strip().upper()
			mes = match.group(2).zfill(2)
			ano = match.group(3)
			
			# Remove acentos e caracteres especiais do município
			municipio = re.sub(r"[^A-Z ]", "", municipio)
			municipio = re.sub(r"\s+", "_", municipio)
			
			return mes, ano, municipio
	except Exception:
		return None


def _extract_table_from_qip_antigo(pdf_path: Path) -> pd.DataFrame | None:
	try:
		with pdfplumber.open(pdf_path) as pdf:
			if not pdf.pages:
				return None
			page = pdf.pages[0]
			tables = page.extract_tables()
			if not tables:
				return None
			
			first_table = tables[0]
			if not first_table:
				return None
			
			# Segunda linha é o cabeçalho (primeira é municipio/mes/ano)
			# Remove a primeira linha (municipio/mes/ano) e a primeira coluna que está vazia
			if len(first_table) < 2:
				return None
			
			headers = [_clean_cell(cell) for cell in first_table[1]]
			data_rows = []
			for row in first_table[2:]:
				cleaned_row = [_clean_cell(cell) for cell in row]
				if any(cleaned_row):
					data_rows.append(cleaned_row)
			
			if not headers or not data_rows:
				return None
			
			df = pd.DataFrame(data_rows, columns=headers)
			return df
	except Exception:
		return None

	name = base if not extra else f"{base}_{extra}"
	name = name[:31]
	idx = 2
	while name in used_names:
		name = f"{base}_{idx}"
		name = name[:31]
		idx += 1
	used_names.add(name)
	return name


def _merge_tables_for_pdf(tables: list[pd.DataFrame]) -> pd.DataFrame:
	if not tables:
		return pd.DataFrame()

	max_columns = max(len(table.columns) for table in tables)
	base_headers = list(tables[0].columns)
	if len(base_headers) < max_columns:
		for index in range(len(base_headers), max_columns):
			base_headers.append(f"coluna_{index + 1}")

	merged_tables: list[pd.DataFrame] = []
	for table in tables:
		rows = table.astype(str).fillna("").values.tolist()
		normalized_rows = [row + [""] * (max_columns - len(row)) for row in rows]
		normalized_table = pd.DataFrame(normalized_rows, columns=base_headers)
		merged_tables.append(normalized_table)

	if not merged_tables:
		return pd.DataFrame(columns=base_headers)

	merged = pd.concat(merged_tables, ignore_index=True)
	return merged


def _apply_worksheet_style(worksheet) -> None:
	from openpyxl.utils import get_column_letter
	import re

	header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
	header_font = Font(color="FFFFFF", bold=True)
	data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
	thin_border = Border(
		left=Side(style="thin", color="D9D9D9"),
		right=Side(style="thin", color="D9D9D9"),
		top=Side(style="thin", color="D9D9D9"),
		bottom=Side(style="thin", color="D9D9D9"),
	)

	max_row = worksheet.max_row
	max_col = worksheet.max_column
	if max_row < 1 or max_col < 1:
		return

	# Detecta tipo de aba: Antigo (primeira célula da primeira linha é tipo texto com mês/ano) ou Novo (cabeçalho padrão)
	first_row = [worksheet.cell(row=1, column=col).value for col in range(1, max_col+1)]
	aba_tipo = "novo"
	if first_row[0]:
		if re.match(r"^[A-Z_]+_\d{2}_\d{4}$", worksheet.title):
			# nome da aba já padronizado
			pass
		# Heurística: se primeira célula da primeira linha é tipo da lâmpada, é antigo
		if str(first_row[0]).strip().lower() in ["tipo da lâmpada", "tipo da lampada"]:
			aba_tipo = "antigo"
		# Se segunda coluna é "Código Lâmpada" ou similar, também é antigo
		if max_col > 1 and str(first_row[1]).strip().lower() in ["código lâmpada", "codigo lampada"]:
			aba_tipo = "antigo"

	# 6. Para QIP novo: encontra a linha "Código da lâmpada"
	# e desloca tudo que está acima dela 8 colunas para a direita.
	if aba_tipo == "novo" and max_row >= 1 and max_col >= 1:
		header_row = None
		for row_index in range(1, max_row + 1):
			for col_index in range(1, max_col + 1):
				cell_value = worksheet.cell(row=row_index, column=col_index).value
				if cell_value is None:
					continue

				normalized = str(cell_value).strip().lower().replace(" ", "")
				if normalized in {"códigodalâmpada", "codigodalampada", "códigolâmpada", "codigolampada"}:
					header_row = row_index
					break
			if header_row is not None:
				break

		if header_row is not None and header_row > 1:
			current_max_col = worksheet.max_column
			for row_index in range(1, header_row):
				for col_index in range(current_max_col, 0, -1):
					origin_cell = worksheet.cell(row=row_index, column=col_index)
					destination_cell = worksheet.cell(row=row_index, column=col_index + 8)
					destination_cell.value = origin_cell.value
					origin_cell.value = None

			# Move a partir da linha do cabeçalho para cima, limitado a A:H,
			# até que a linha "Código da lâmpada" vire a linha 1.
			last_col = min(8, worksheet.max_column)
			shift_up = header_row - 1
			for row_index in range(header_row, max_row + 1):
				for col_index in range(1, last_col + 1):
					worksheet.cell(row=row_index - shift_up, column=col_index).value = worksheet.cell(
						row=row_index,
						column=col_index,
					).value

			# Limpa as últimas linhas que ficaram duplicadas após o deslocamento para cima (somente A:H).
			for row_index in range(max_row - shift_up + 1, max_row + 1):
				for col_index in range(1, last_col + 1):
					worksheet.cell(row=row_index, column=col_index).value = None

	# Recalcula limites após possíveis alterações de layout
	max_row = worksheet.max_row
	max_col = worksheet.max_column

	worksheet.freeze_panes = "A2"
	worksheet.auto_filter.ref = worksheet.dimensions

	# 1. Converter células numéricas para tipo numérico real
	num_regex = re.compile(r"^-?\d[\d\s\.,]*$")
	for row_index in range(2, max_row + 1):
		for col_index in range(1, max_col + 1):
			cell = worksheet.cell(row=row_index, column=col_index)
			val = cell.value
			if isinstance(val, str):
				val_limpo = val.strip()
				if not val_limpo or not num_regex.match(val_limpo):
					continue

				val_sem_espaco = val_limpo.replace(" ", "")
				if aba_tipo == "antigo":
					# Regra solicitada para antigo: remove vírgulas e troca ponto por vírgula
					val_corrigido = val_sem_espaco.replace(",", "").replace(".", ",")
					normalizado = val_corrigido.replace(",", ".")
				else:
					# Para novo, só normaliza separador sem mexer em textos
					if "," in val_sem_espaco and "." in val_sem_espaco:
						normalizado = val_sem_espaco.replace(".", "").replace(",", ".")
					elif "," in val_sem_espaco:
						normalizado = val_sem_espaco.replace(",", ".")
					else:
						normalizado = val_sem_espaco

				try:
					if "." in normalizado:
						cell.value = float(normalizado)
					else:
						cell.value = int(normalizado)
				except Exception:
					pass

	# 2. Centralizar horizontalmente células numéricas, 3. Centralizar verticalmente todas
	for row_index in range(1, max_row + 1):
		for col_index in range(1, max_col + 1):
			cell = worksheet.cell(row=row_index, column=col_index)
			val = cell.value
			if isinstance(val, (int, float)):
				cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			else:
				cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
			cell.border = thin_border
			if row_index == 1:
				cell.fill = header_fill
				cell.font = header_font
			elif row_index % 2 == 0:
				cell.fill = data_fill

	# 4. Primeira linha altura 30
	worksheet.row_dimensions[1].height = 30

	# 5. Largura das colunas
	if aba_tipo == "antigo":
		for col_index in range(1, max_col + 1):
			col_letter = get_column_letter(col_index)
			if col_index == 1:
				worksheet.column_dimensions[col_letter].width = 30
			else:
				worksheet.column_dimensions[col_letter].width = 15
	elif aba_tipo == "novo":
		for col_index in range(1, max_col + 1):
			col_letter = get_column_letter(col_index)
			if col_index == 2:
				worksheet.column_dimensions[col_letter].width = 30
			else:
				worksheet.column_dimensions[col_letter].width = 15


def extract_tables_from_pdf(pdf_path: Path) -> list[pd.DataFrame]:
	tables: list[pd.DataFrame] = []

	with pdfplumber.open(pdf_path) as pdf:
		for page in pdf.pages:
			page_tables = page.extract_tables() or []
			for raw_table in page_tables:
				if not raw_table:
					continue

				dataframe = _table_to_dataframe(raw_table)
				if not dataframe.empty:
					tables.append(dataframe)

	return tables


def save_tables_to_excel(tables: list[pd.DataFrame], output_file: Path) -> None:
	output_file.parent.mkdir(parents=True, exist_ok=True)

	with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
		merged_table = _merge_tables_for_pdf(tables)
		sheet_name = "QIP"
		merged_table.to_excel(writer, sheet_name=sheet_name, index=False)
		_apply_worksheet_style(writer.sheets[sheet_name])


def save_pdfs_to_single_workbook(pdf_dataframes: dict[str, pd.DataFrame], output_file: Path) -> None:
	output_file.parent.mkdir(parents=True, exist_ok=True)
	used_sheet_names: set[str] = set()

	with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
		for pdf_name, dataframe in pdf_dataframes.items():
			sheet_name = _sanitize_sheet_name(pdf_name, used_sheet_names)
			dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
			_apply_worksheet_style(writer.sheets[sheet_name])


def convert_pdf_to_xlsx(pdf_file: Path, output_dir: Path) -> Path:
	tables = extract_tables_from_pdf(pdf_file)
	if not tables:
		raise ValueError(
			f"Nenhuma tabela foi encontrada no arquivo PDF: {pdf_file.name}"
		)

	output_file = output_dir / f"{pdf_file.stem}.xlsx"
	save_tables_to_excel(tables, output_file)
	return output_file


def process_folder(test_folder: Path, output_dir: Path) -> Path:
	pdf_files = sorted(test_folder.glob("*.pdf"))
	if not pdf_files:
		raise FileNotFoundError(
			f"Nenhum PDF encontrado na pasta de testes: {test_folder}"
		)

	pdf_dataframes: dict[str, pd.DataFrame] = {}
	total_pdfs = len(pdf_files)
	print(f"Iniciando processamento de {total_pdfs} PDF(s)...")
	used_sheet_names: set[str] = set()

	def process_single_pdf(pdf_file):
		try:
			is_antigo = _is_qip_antigo(pdf_file)
			if is_antigo:
				info = _extract_mes_ano_municipio_from_qip_antigo(pdf_file)
				if not info:
					return (pdf_file, None, f"[AVISO] Não foi possível extrair dados do QIP antigo '{pdf_file.name}', pulando...")
				mes, ano, municipio = info
				novo_nome = f"QIP-{municipio}-{ano}_{mes}.pdf"
				novo_pdf_path = pdf_file.parent / novo_nome
				if pdf_file.name != novo_nome:
					if not novo_pdf_path.exists():
						pdf_file.rename(novo_pdf_path)
					else:
						novo_pdf_path = pdf_file
				else:
					novo_pdf_path = pdf_file
				table = _extract_table_from_qip_antigo(novo_pdf_path)
				if table is None or table.empty:
					return (pdf_file, None, f"[AVISO] Não foi possível extrair tabela do QIP antigo '{novo_pdf_path.name}', pulando...")
				aba_nome = f"{ano}_{mes}_{municipio}"
				return (aba_nome, table, None)
			else:
				info = _extract_mes_ano_municipio_from_pdf(pdf_file)
				if not info:
					return (pdf_file, None, f"[AVISO] Não foi possível extrair mês/ano/município do conteúdo de '{pdf_file.name}', pulando...")
				mes, ano, municipio = info
				novo_nome = f"QIP-{municipio}-{ano}_{mes}.pdf"
				novo_pdf_path = pdf_file.parent / novo_nome
				if pdf_file.name != novo_nome:
					if not novo_pdf_path.exists():
						pdf_file.rename(novo_pdf_path)
					else:
						novo_pdf_path = pdf_file
				else:
					novo_pdf_path = pdf_file
				tables = extract_tables_from_pdf(novo_pdf_path)
				if not tables:
					return (pdf_file, None, f"[AVISO] Não foi possível extrair tabelas do QIP Novo: {novo_pdf_path.name}, pulando...")
				merged_table = _merge_tables_for_pdf(tables)
				aba_nome = f"{ano}_{mes}_{municipio}"
				return (aba_nome, merged_table, None)
		except Exception as e:
			return (pdf_file, None, f"[ERRO] Falha inesperada em {pdf_file.name}: {e}")

	results = []
	with ThreadPoolExecutor() as executor:
		future_to_pdf = {executor.submit(process_single_pdf, pdf_file): pdf_file for pdf_file in pdf_files}
		for idx, future in enumerate(as_completed(future_to_pdf), 1):
			aba_nome, dataframe, error = future.result()
			if error:
				print(error)
				continue
			aba_nome = _sanitize_sheet_name(str(aba_nome), used_sheet_names)
			pdf_dataframes[aba_nome] = dataframe
			_render_progress(idx, total_pdfs, prefix="Processando PDFs")

	if not pdf_dataframes:
		raise ValueError("Nenhuma tabela válida foi extraída dos PDFs encontrados.")

	output_file = output_dir / DEFAULT_WORKBOOK_NAME
	save_pdfs_to_single_workbook(pdf_dataframes, output_file)
	return output_file

def save_pdfs_to_single_workbook(pdf_dataframes: dict[str, pd.DataFrame], output_file: Path) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    used_sheet_names: set[str] = set()

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet_name, dataframe in pdf_dataframes.items():
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
            _apply_worksheet_style(writer.sheets[sheet_name])


def build_parser() -> argparse.ArgumentParser:
	parser = argparse.ArgumentParser(
		description="Converte tabelas de PDFs de QIP para arquivos XLSX."
	)
	parser.add_argument(
		"--pdf",
		type=Path,
		help="Caminho de um PDF específico para converter.",
	)
	parser.add_argument(
		"--testes",
		type=Path,
		default=TESTES_DIR,
		help="Pasta onde estão os PDFs de teste (padrão: Organizador_de_QIP/Testes).",
	)
	parser.add_argument(
		"--saida",
		type=Path,
		default=OUTPUT_DIR,
		help="Pasta de saída dos arquivos XLSX.",
	)
	parser.add_argument(
		"--arquivo-saida",
		type=str,
		default=DEFAULT_WORKBOOK_NAME,
		help="Nome do arquivo XLSX gerado quando processar pasta de testes.",
	)
	return parser


def main() -> None:
	parser = build_parser()
	args = parser.parse_args()

	if args.pdf:
		pdf_file = args.pdf.resolve()
		output_file = convert_pdf_to_xlsx(pdf_file, args.saida.resolve())
		print(f"Arquivo gerado: {output_file}")
		return

	output_dir = args.saida.resolve()
	generated = process_folder(args.testes.resolve(), output_dir)

	if args.arquivo_saida != DEFAULT_WORKBOOK_NAME:
		custom_output = output_dir / args.arquivo_saida
		generated.replace(custom_output)
		generated = custom_output

	print("Arquivo consolidado gerado:")
	print(f"- {generated}")


if __name__ == "__main__":
	main()
