from pathlib import Path
from datetime import datetime, date
from calendar import monthrange
from collections import Counter, defaultdict
import json
import os
import re
import shutil
import subprocess
import sys
import unicodedata
import urllib.error
import urllib.request
import warnings

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

warnings.simplefilter("ignore")

ROOT = Path(__file__).resolve().parent
OUTPUTS = ROOT / "outputs"
OUTPUTS.mkdir(parents=True, exist_ok=True)

ONEDRIVE_BASE = Path(r"C:\Users\Usuário 1\Documents\Inovve_Automação\APURAÇÃO")
FDOJ_CE_FILE = ONEDRIVE_BASE / "FDOJ - Ceará" / "FDOJ RECLAMAÇÕES.xlsm"
HLA_CE_FILE = ONEDRIVE_BASE / "Municípios HLA - Ceará" / "CONTROLE DE RECLAMAÇÕES - CEARÁ HLA.xlsm"
GRID_CE_FILE = ONEDRIVE_BASE / "Anderson" / "Acompanhamento diario" / "GRID - controle processos reclamações.xlsm"
INOVVE_CE_FILE = ONEDRIVE_BASE / "Anderson" / "Acompanhamento diario" / "INOVVE - controle processos reclamações ceara - Editável - Editável.xlsm"
FDOJ_OUTROS_FILE = ONEDRIVE_BASE / "FDOJ - Outros Estados" / "CONTROLE DE RECLAMAÇÕES - FDOJ.xlsm"
GRID_OUTROS_FILE = ONEDRIVE_BASE / "Municípios Grid - Outros Estados" / "CONTROLE DE RECLAMAÇÕES - GRID.xlsm"
HLA_OUTROS_FILE = ONEDRIVE_BASE / "Municípios HLA - Outros Estados" / "CONTROLE DE RECLAMAÇÕES - HLA.xlsm"
INOVVE_OUTROS_FILE = ONEDRIVE_BASE / "Municípios Inovve - Outros Estados" / "CONTROLE DE RECLAMAÇÕES - INOVVE ESTADOS DE FORA 1.xlsm"
ABEL_FILE = ONEDRIVE_BASE / "PARCEIROS" / "Municípios Abel Gomes Cunha" / "CONTROLE DE RECLAMAÇÕES - ABEL CUNHA-HLA-ENERGIA-06.xlsm"
AUGUSTO_FILE = ONEDRIVE_BASE / "PARCEIROS" / "Municípios Augusto Santos" / "CONTROLE DE RECLAMAÇÕES - AUGUSTO SANTOS.xlsm"
INDYRA_FILE = ONEDRIVE_BASE / "PARCEIROS" / "Municípios Indyra" / "CONTROLE DE RECLAMAÇÕES - INDYRA.xlsm"
MONTEIRO_FILE = ONEDRIVE_BASE / "PARCEIROS" / "Municípios Monteiro & Monteiro" / "Estados" / "Planilhas de Controle" / "CONTROLE DE RECLAMAÇÕES - MONTEIRO E MONTEIRO.xlsm"
RUDA_FILE = ONEDRIVE_BASE / "CONTROLE DE RECLAMAÇÕES - THAMIRES E RUDÁ-HLA-ENERGIA-06.xlsm"
OLIVEIRA_VARELA_FILE = ONEDRIVE_BASE / "PARCEIROS" / "Municípios Wagner (Oliveira e Varela - Advogados)" / "CONTROLE DE RECLAMAÇÕES - OLIVEIRA E VARELA.xlsm"
PRIVADOS_FILE = ONEDRIVE_BASE / "CLIENTES PRIVADOS" / "CLIENTES PRIVADOS - CONTROLE" / "CONTROLE DE RECLAMAÇÕES - PRIVADOS (USAR ESSE).xlsm"
NF_FILE = ONEDRIVE_BASE / "COORDENAÇÃO" / "CONTROLE" / "CONTROLE NF" / "PROCESSO ENVIADOS PRA GERAR NF.xlsx"
WEB_TASKS_URL = os.environ.get("TASK_WEB_URL", "https://controle-equipe.onrender.com").rstrip("/")
EMPRESAS_RECLAMACOES = ["FDOJ", "HLA", "GRID", "INOVVE", "ABEL", "AUGUSTO", "INDYRA", "MONTEIRO", "RUDÁ", "OLIVEIRA E VARELA", "PRIVADOS"]

STANDARD_SOURCE_COLUMNS = {
    "sheet": "ANDAMENTO TESES",
    "header_row": 5,
    "rec_req": 5,
    "tese": 6,
    "concessionaria": {"responsavel": 16, "data": 17, "municipio": 3},
    "ouvidoria": {"responsavel": 24, "data": 25, "municipio": 3},
    "aneel": {"responsavel": 33, "data": 34, "municipio": 3},
    "processo_adm": {"responsavel": 40, "data": 41, "municipio": 3},
}

def make_standard_source(empresa, path, processo_adm=True):
    source = {"empresa": empresa, "path": path, **STANDARD_SOURCE_COLUMNS}
    if not processo_adm:
        source["processo_adm"] = None
    return source

INGRESSOS_SOURCES = [
    {
        "empresa": "FDOJ",
        "path": FDOJ_CE_FILE,
        "sheet": "ANDAMENTO TESES",
        "header_row": 2,
        "rec_req": 7,
        "tese": 9,
        "ouvidoria": {"responsavel": 26, "data": 27, "municipio": 5},
        "processo_adm": {"responsavel": 42, "data": 43, "municipio": 5},
    },
    {
        "empresa": "HLA",
        "path": HLA_CE_FILE,
        "sheet": "ANDAMENTO TESES",
        "header_row": 5,
        "rec_req": 5,
        "tese": 6,
        "ouvidoria": {"responsavel": 24, "data": 25, "municipio": 3},
        "processo_adm": None,
    },
    {
        "empresa": "GRID",
        "path": GRID_CE_FILE,
        "sheet": "ANDAMENTO TESES",
        "header_row": 2,
        "rec_req": 5,
        "tese": 6,
        "ouvidoria": {"responsavel": 21, "data": 22, "municipio": 3},
        "processo_adm": {"responsavel": 37, "data": 38, "municipio": 3},
    },
    {
        "empresa": "INOVVE",
        "path": INOVVE_CE_FILE,
        "sheet": "ANDAMENTO TESES",
        "header_row": 2,
        "rec_req": 6,
        "tese": 7,
        "ouvidoria": {"responsavel": 24, "data": 25, "municipio": 4},
        "processo_adm": {"responsavel": 40, "data": 41, "municipio": 4},
    },
    {
        "empresa": "FDOJ",
        "path": FDOJ_OUTROS_FILE,
        "sheet": "Controle",
        "header_row": 5,
        "rec_req": 5,
        "tese": 6,
        "concessionaria": {"responsavel": 16, "data": 17, "municipio": 3},
        "ouvidoria": {"responsavel": 24, "data": 25, "municipio": 3},
        "aneel": {"responsavel": 33, "data": 34, "municipio": 3},
        "processo_adm": None,
    },
    make_standard_source("GRID", GRID_OUTROS_FILE, processo_adm=False),
    make_standard_source("HLA", HLA_OUTROS_FILE),
    make_standard_source("INOVVE", INOVVE_OUTROS_FILE, processo_adm=False),
    make_standard_source("ABEL", ABEL_FILE),
    make_standard_source("AUGUSTO", AUGUSTO_FILE),
    make_standard_source("INDYRA", INDYRA_FILE),
    make_standard_source("MONTEIRO", MONTEIRO_FILE),
    make_standard_source("RUDÁ", RUDA_FILE),
    make_standard_source("OLIVEIRA E VARELA", OLIVEIRA_VARELA_FILE),
    {
        "empresa": "PRIVADOS",
        "path": PRIVADOS_FILE,
        "sheet": "ANDAMENTO TESES",
        "header_row": 5,
        "rec_req": 5,
        "tese": 6,
        "concessionaria": {"responsavel": 15, "data": 16, "municipio": 3},
        "ouvidoria": {"responsavel": 23, "data": 24, "municipio": 3},
        "aneel": {"responsavel": 32, "data": 33, "municipio": 3},
        "processo_adm": {"responsavel": 39, "data": 40, "municipio": 3},
    },
]

TAREFAS_FILES = []
TAREFAS_WEB_START = date(9999, 1, 1)

DEFERIMENTOS_SOURCES = [
    {"empresa": "FDOJ", "path": FDOJ_CE_FILE, "sheet": "ANDAMENTO TESES", "header_row": 2, "municipio": 5, "data_deferimento": 13, "valor_deferido": 14, "reclamacao": 7, "tese": 9},
    {"empresa": "HLA", "path": HLA_CE_FILE, "sheet": "ANDAMENTO TESES", "header_row": 5, "municipio": 3, "data_deferimento": 10, "valor_deferido": 11, "reclamacao": 5, "tese": 6},
    {"empresa": "GRID", "path": GRID_CE_FILE, "sheet": "ANDAMENTO TESES", "header_row": 2, "municipio": 3, "data_deferimento": 8, "valor_deferido": 9, "reclamacao": 5, "tese": 6},
    {"empresa": "INOVVE", "path": INOVVE_CE_FILE, "sheet": "ANDAMENTO TESES", "header_row": 2, "municipio": 4, "data_deferimento": 11, "valor_deferido": 12, "reclamacao": 6, "tese": 7},
    {"empresa": "FDOJ", "path": FDOJ_OUTROS_FILE, "sheet": "Controle", "header_row": 5, "municipio": 3, "data_deferimento": 10, "valor_deferido": 12, "reclamacao": 5, "tese": 6},
    {"empresa": "GRID", "path": GRID_OUTROS_FILE, "sheet": "ANDAMENTO TESES", "header_row": 5, "municipio": 3, "data_deferimento": 10, "valor_deferido": 11, "reclamacao": 5, "tese": 6},
    {"empresa": "HLA", "path": HLA_OUTROS_FILE, "sheet": "ANDAMENTO TESES", "header_row": 5, "municipio": 3, "data_deferimento": 10, "valor_deferido": 12, "reclamacao": 5, "tese": 6},
    {"empresa": "INOVVE", "path": INOVVE_OUTROS_FILE, "sheet": "ANDAMENTO TESES", "header_row": 5, "municipio": 3, "data_deferimento": 10, "valor_deferido": 12, "reclamacao": 5, "tese": 6},
    {"empresa": "ABEL", "path": ABEL_FILE, "sheet": "ANDAMENTO TESES", "header_row": 5, "municipio": 3, "data_deferimento": 10, "valor_deferido": 12, "reclamacao": 5, "tese": 6},
    {"empresa": "AUGUSTO", "path": AUGUSTO_FILE, "sheet": "ANDAMENTO TESES", "header_row": 5, "municipio": 3, "data_deferimento": 10, "valor_deferido": 12, "reclamacao": 5, "tese": 6},
    {"empresa": "INDYRA", "path": INDYRA_FILE, "sheet": "ANDAMENTO TESES", "header_row": 5, "municipio": 3, "data_deferimento": 10, "valor_deferido": 12, "reclamacao": 5, "tese": 6},
    {"empresa": "MONTEIRO", "path": MONTEIRO_FILE, "sheet": "ANDAMENTO TESES", "header_row": 5, "municipio": 3, "data_deferimento": 10, "valor_deferido": 12, "reclamacao": 5, "tese": 6},
    {"empresa": "RUDÁ", "path": RUDA_FILE, "sheet": "ANDAMENTO TESES", "header_row": 5, "municipio": 3, "data_deferimento": 10, "valor_deferido": 12, "reclamacao": 5, "tese": 6},
    {"empresa": "OLIVEIRA E VARELA", "path": OLIVEIRA_VARELA_FILE, "sheet": "ANDAMENTO TESES", "header_row": 5, "municipio": 3, "data_deferimento": 10, "valor_deferido": 11, "reclamacao": 5, "tese": 6},
    {"empresa": "PRIVADOS", "path": PRIVADOS_FILE, "sheet": "ANDAMENTO TESES", "header_row": 5, "municipio": 3, "data_deferimento": 10, "valor_deferido": 11, "reclamacao": 5, "tese": 6},
]
CACHE_DIR = ROOT / "work" / "_apuracao_cache"


def parse_period(arg):
    value = arg.strip()
    match = re.fullmatch(r"(\d{1,2})/(\d{4})", value)
    if not match:
        raise ValueError("Informe o período no formato MM/AAAA, por exemplo 05/2026.")
    month = int(match.group(1))
    year = int(match.group(2))
    if month < 1 or month > 12:
        raise ValueError("Mês inválido.")
    start = date(year, month, 1)
    end = date(year, month, monthrange(year, month)[1])
    return month, year, start, end


def parse_interval_date(arg):
    value = arg.strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass
    raise ValueError("Informe a data no formato DD/MM/AAAA, por exemplo 01/05/2026.")


def load_workbook_safe(path, **kwargs):
    try:
        return load_workbook(path, **kwargs)
    except PermissionError:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", path.name)
        cached = CACHE_DIR / safe_name
        shutil.copy2(path, cached)
        return load_workbook(cached, **kwargs)


def clean_text(value, default=""):
    if value is None:
        return default
    text = re.sub(r"\s+", " ", str(value).strip())
    return text if text else default


def clean_name(value):
    text = clean_text(value, "S/I")
    key = text.lower()
    if key in {"s/i", "si", "n/a", "na", "-", "0"}:
        return "S/I"
    if "anderson" in key:
        return "Anderson"
    if "debora" in key or "débora" in key:
        return "Débora"
    if "nickolas" in key:
        return "Nickolas"
    if "gonzaga" in key:
        return "Gonzaga"
    if "larissa" in key:
        return "Larissa"
    if "mayara" in key:
        return "Mayara"
    return text.title()


def clean_city(value):
    text = clean_text(value, "S/I")
    if text.upper() in {"S/I", "SI", "N/A", "NA", "-", "0"}:
        return "S/I"
    return text.upper()


def clean_state(value):
    text = clean_text(value, "S/I")
    if text.upper() in {"S/I", "SI", "N/A", "NA", "-", "0"}:
        return "S/I"
    return text.upper()


def city_key(value):
    text = clean_city(value)
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", normalized).strip().upper()


def allowed_city(source, municipio):
    allowed = source.get("municipios_permitidos")
    if not allowed:
        return True
    return city_key(municipio) in {city_key(item) for item in allowed}


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text or text.upper() in {"S/I", "SI", "N/A", "NA", "-", "?"}:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%dT%H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text or text.lower() in {"sem informação", "sem informacao", "s/i", "-", "na", "n/a"}:
        return None
    text = text.replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def parse_money(value):
    parsed = parse_number(value)
    return parsed if parsed is not None else 0.0


def parse_rate(value):
    parsed = parse_number(value)
    if parsed is None:
        return None
    return parsed / 100 if parsed > 1 else parsed


def should_report_invalid_date(value, include_question=False):
    if value is None or isinstance(value, (date, datetime, int, float)):
        return False
    text = clean_text(value)
    if not text or text.upper() in {"S/I", "SI", "N/A", "NA", "-"}:
        return False
    if include_question and text == "?":
        return True
    return any(ch.isdigit() for ch in text)


def pct(part, total):
    return part / total if total else 0


def pct100(part, total):
    return round(pct(part, total) * 100, 2)


def previous_period(start):
    prev_month = 12 if start.month == 1 else start.month - 1
    prev_year = start.year - 1 if start.month == 1 else start.year
    prev_start = date(prev_year, prev_month, 1)
    prev_end = date(prev_year, prev_month, monthrange(prev_year, prev_month)[1])
    return prev_month, prev_year, prev_start, prev_end


def compare_metric(grupo, indicador, atual, anterior, unidade="Quantidade"):
    variacao = atual - anterior
    variacao_pct = variacao / anterior if anterior else (1 if atual else 0)
    if variacao > 0:
        status = "Aumentou"
    elif variacao < 0:
        status = "Diminuiu"
    else:
        status = "Estável"
    return {
        "Grupo": grupo,
        "Indicador": indicador,
        "Atual": atual,
        "Mês Anterior": anterior,
        "Variação": variacao,
        "% Variação": variacao_pct,
        "Status": status,
        "Unidade": unidade,
    }


def build_comparativo(atual, anterior):
    ingressos_atual = atual["ingressos"]
    ingressos_ant = anterior["ingressos"]
    tarefas_atual = atual["tarefas"]
    tarefas_ant = anterior["tarefas"]
    nf_atual = atual["nf"]
    nf_ant = anterior["nf"]
    def_atual = atual["deferimentos"]
    def_ant = anterior["deferimentos"]

    rows = [
        compare_metric("Ingressos", "Total de ingressos", ingressos_atual["totais"]["Total"], ingressos_ant["totais"]["Total"]),
        compare_metric("Ingressos", "Concessionária", ingressos_atual["totais"].get("Concessionária", 0), ingressos_ant["totais"].get("Concessionária", 0)),
        compare_metric("Ingressos", "Ouvidoria", ingressos_atual["totais"]["Ouvidoria"], ingressos_ant["totais"]["Ouvidoria"]),
        compare_metric("Ingressos", "Ouvidoria ANEEL", ingressos_atual["totais"].get("Ouvidoria ANEEL", 0), ingressos_ant["totais"].get("Ouvidoria ANEEL", 0)),
        compare_metric("Ingressos", "Processo ADM", ingressos_atual["totais"]["Processo ADM"], ingressos_ant["totais"]["Processo ADM"]),
        compare_metric("Atividades", "Total de atividades", tarefas_atual["total"], tarefas_ant["total"]),
        compare_metric("Faturamento NF", "Quantidade Anderson", nf_atual["totais"]["Quantidade Anderson"], nf_ant["totais"]["Quantidade Anderson"]),
        compare_metric("Faturamento NF", "Serviço/retorno Anderson", nf_atual["totais"]["Valor Anderson"], nf_ant["totais"]["Valor Anderson"], "Valor"),
        compare_metric("Faturamento NF", "Valor dos processos Anderson", nf_atual["totais"]["Valor Processos Anderson"], nf_ant["totais"]["Valor Processos Anderson"], "Valor"),
        compare_metric("Deferimentos", "Quantidade de deferimentos", def_atual["totais"]["Quantidade"], def_ant["totais"]["Quantidade"]),
        compare_metric("Deferimentos", "Valor deferido", def_atual["totais"]["Valor Deferido"], def_ant["totais"]["Valor Deferido"], "Valor"),
    ]

    atual_emp = {r["Empresa"]: r for r in ingressos_atual["resumo_empresa"]}
    ant_emp = {r["Empresa"]: r for r in ingressos_ant["resumo_empresa"]}
    for empresa in EMPRESAS_RECLAMACOES:
        rows.append(compare_metric("Ingressos por empresa", empresa, atual_emp.get(empresa, {}).get("Total", 0), ant_emp.get(empresa, {}).get("Total", 0)))

    atual_resp_ing = {r["Responsável"]: r for r in ingressos_atual["responsaveis"]}
    ant_resp_ing = {r["Responsável"]: r for r in ingressos_ant["responsaveis"]}
    for resp in sorted(set(atual_resp_ing) | set(ant_resp_ing)):
        rows.append(compare_metric("Ingressos por colaborador", resp, atual_resp_ing.get(resp, {}).get("Total", 0), ant_resp_ing.get(resp, {}).get("Total", 0)))

    atual_ativ = {r["Analista"]: r for r in tarefas_atual["resumo"]}
    ant_ativ = {r["Analista"]: r for r in tarefas_ant["resumo"]}
    for analista in sorted(set(atual_ativ) | set(ant_ativ)):
        rows.append(compare_metric("Atividades por analista", analista, atual_ativ.get(analista, {}).get("Total de Atividades", 0), ant_ativ.get(analista, {}).get("Total de Atividades", 0)))

    atual_nf_emp = {r["Empresa"]: r for r in nf_atual["por_empresa_anderson"]}
    ant_nf_emp = {r["Empresa"]: r for r in nf_ant["por_empresa_anderson"]}
    for empresa in ["INOVVE", "HLA", "GRID", "FDOJ"]:
        rows.append(compare_metric("NF Anderson por empresa", empresa, atual_nf_emp.get(empresa, {}).get("Valor Total", 0), ant_nf_emp.get(empresa, {}).get("Valor Total", 0), "Valor"))

    atual_def_emp = {r["Empresa"]: r for r in def_atual["por_empresa"]}
    ant_def_emp = {r["Empresa"]: r for r in def_ant["por_empresa"]}
    for empresa in EMPRESAS_RECLAMACOES:
        rows.append(compare_metric("Deferimentos por empresa", empresa, atual_def_emp.get(empresa, {}).get("Valor Deferido", 0), ant_def_emp.get(empresa, {}).get("Valor Deferido", 0), "Valor"))

    return rows


def company_from_contract(value):
    text = clean_text(value).upper()
    for company in ("INOVVE", "HLA", "GRID", "FDOJ"):
        if company in text:
            return company
    return text or "S/I"


def norm_header(value):
    text = clean_text(value).upper()
    accents = str.maketrans("ÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇ", "AAAAAEEEEIIIIOOOOOUUUUC")
    return text.translate(accents)


def analista_from_file(path):
    name = path.stem.lower()
    if "anderson" in name:
        return "Anderson"
    if "debora" in name or "débora" in name:
        return "Débora"
    if "nickolas" in name:
        return "Nickolas"
    if "gonzaga" in name:
        return "Gonzaga"
    return path.stem


def apurar_ingressos(start, end):
    records = []
    inconsistencias = []
    for source in INGRESSOS_SOURCES:
        if not source["path"].exists():
            inconsistencias.append({
                "Origem": "Ingressos",
                "Arquivo": source["path"].name,
                "Aba": source["sheet"],
                "Linha": "",
                "Tipo": "Arquivo não encontrado",
                "Detalhe": str(source["path"]),
            })
            continue
        wb = load_workbook_safe(source["path"], read_only=True, data_only=True, keep_vba=True)
        if source["sheet"] not in wb.sheetnames:
            inconsistencias.append({
                "Origem": "Ingressos",
                "Arquivo": source["path"].name,
                "Aba": source["sheet"],
                "Linha": "",
                "Tipo": "Aba não encontrada",
                "Detalhe": f"Abas disponíveis: {', '.join(wb.sheetnames)}",
            })
            continue
        ws = wb[source["sheet"]]
        for row_num, row in enumerate(ws.iter_rows(min_row=source["header_row"] + 1, values_only=False), start=source["header_row"] + 1):
            levels = [
                ("Concessionária", source.get("concessionaria")),
                ("Ouvidoria", source.get("ouvidoria")),
                ("Ouvidoria ANEEL", source.get("aneel")),
                ("Processo ADM", source.get("processo_adm")),
            ]
            municipio_cols = [cols["municipio"] for _, cols in levels if cols is not None]
            row_municipio = clean_city(row[municipio_cols[0] - 1].value) if municipio_cols else "S/I"
            row_estado = clean_state(row[municipio_cols[0]].value) if municipio_cols and len(row) > municipio_cols[0] else "S/I"
            if not allowed_city(source, row_municipio):
                continue
            for tipo, cols in levels:
                if cols is None:
                    continue
                raw_date = row[cols["data"] - 1].value
                date_value = parse_date(raw_date)
                if should_report_invalid_date(raw_date) and date_value is None:
                    inconsistencias.append({
                        "Origem": "Ingressos",
                        "Arquivo": source["path"].name,
                        "Aba": source["sheet"],
                        "Linha": row_num,
                        "Tipo": f"Data inválida em {tipo}",
                        "Detalhe": clean_text(raw_date),
                    })
                    continue
                if date_value is None or date_value < start or date_value > end:
                    continue
                if cols.get("responsavel"):
                    responsavel = clean_name(row[cols["responsavel"] - 1].value)
                else:
                    responsavel = clean_name(cols.get("responsavel_default", "S/I"))
                records.append({
                    "Empresa": source["empresa"],
                    "Tipo de Ingresso": tipo,
                    "Nº REC./REQ.": clean_text(row[source["rec_req"] - 1].value),
                    "Tese": clean_text(row[source["tese"] - 1].value),
                    "Responsável": responsavel,
                    "Parceiro": clean_text(row[0].value, "S/I"),
                    "Município": row_municipio,
                    "Estado": row_estado,
                    "Data": date_value.isoformat(),
                    "Linha Original": row_num,
                    "Arquivo": source["path"].name,
                })
    ouvidoria = [r for r in records if r["Tipo de Ingresso"] == "Ouvidoria"]
    concessionaria = [r for r in records if r["Tipo de Ingresso"] == "Concessionária"]
    aneel = [r for r in records if r["Tipo de Ingresso"] == "Ouvidoria ANEEL"]
    processo = [r for r in records if r["Tipo de Ingresso"] == "Processo ADM"]
    total = len(records)
    resumo_empresa = []
    for empresa in EMPRESAS_RECLAMACOES:
        total_conc = sum(1 for r in concessionaria if r["Empresa"] == empresa)
        total_ouv = sum(1 for r in ouvidoria if r["Empresa"] == empresa)
        total_aneel = sum(1 for r in aneel if r["Empresa"] == empresa)
        total_proc = sum(1 for r in processo if r["Empresa"] == empresa)
        resumo_empresa.append({
            "Empresa": empresa,
            "Concessionária": total_conc,
            "Ouvidoria": total_ouv,
            "Ouvidoria ANEEL": total_aneel,
            "Processo ADM": total_proc,
            "Total": total_conc + total_ouv + total_aneel + total_proc,
            "% do Total": pct(total_conc + total_ouv + total_aneel + total_proc, total),
        })
    responsaveis = [
        {"Responsável": resp, "Total": count, "% do Total": pct(count, total)}
        for resp, count in Counter(r["Responsável"] for r in records).most_common()
    ]
    municipios = [
        {"Parceiro": parceiro, "Município": municipio, "Estado": estado, "Total": count, "% do Total": pct(count, total)}
        for (parceiro, municipio, estado), count in Counter((r["Parceiro"], r["Município"], r["Estado"]) for r in records).most_common()
    ]
    return {
        "records": records,
        "resumo_empresa": resumo_empresa,
        "responsaveis": responsaveis,
        "municipios": municipios,
        "totais": {
            "Concessionária": len(concessionaria),
            "Ouvidoria": len(ouvidoria),
            "Ouvidoria ANEEL": len(aneel),
            "Processo ADM": len(processo),
            "Total": total,
        },
        "inconsistencias": inconsistencias,
    }


def retorno_em_aberto(value):
    if value is None:
        return True
    if isinstance(value, (datetime, date)):
        return False
    if isinstance(value, (int, float)) and value == 0:
        return True
    text = clean_text(value).upper()
    if text in {"", "S/I", "SI", "N/A", "NA", "-", "0", "ZERO"}:
        return True
    return parse_date(value) is None


def find_ouvidoria_return_col(ws, header_row, protocolo_col):
    max_col = min(ws.max_column, protocolo_col + 10)
    for row_num in range(max(1, header_row - 2), min(ws.max_row, header_row + 3) + 1):
        for col in range(protocolo_col + 1, max_col + 1):
            header = norm_header(ws.cell(row_num, col).value)
            if "RETORNO" in header:
                return col
    return protocolo_col + 2


def find_ouvidoria_protocol_number_col(ws, header_row, data_col):
    max_col = min(ws.max_column, data_col + 5)
    for row_num in range(max(1, header_row - 2), min(ws.max_row, header_row + 3) + 1):
        for col in range(data_col + 1, max_col + 1):
            header = norm_header(ws.cell(row_num, col).value)
            if "PROTOCOLO" in header and "DATA" not in header:
                return col
    return data_col + 1


def apurar_ouvidoria_abertos(end):
    base_start = date(2025, 1, 1)
    records = []
    inconsistencias = []
    for source in INGRESSOS_SOURCES:
        if not source.get("ouvidoria_abertos_enel_ce"):
            continue
        cols = source.get("ouvidoria")
        if cols is None:
            continue
        wb = load_workbook_safe(source["path"], read_only=True, data_only=True, keep_vba=True)
        ws = wb[source["sheet"]]
        protocolo_ouvidoria_col = find_ouvidoria_protocol_number_col(ws, source["header_row"], cols["data"])
        retorno_col = find_ouvidoria_return_col(ws, source["header_row"], cols["data"])
        for row_num, row in enumerate(ws.iter_rows(min_row=source["header_row"] + 1, values_only=False), start=source["header_row"] + 1):
            raw_protocol = row[cols["data"] - 1].value
            protocol_date = parse_date(raw_protocol)
            if should_report_invalid_date(raw_protocol) and protocol_date is None:
                inconsistencias.append({
                    "Origem": "Ouvidoria em aberto",
                    "Arquivo": source["path"].name,
                    "Aba": source["sheet"],
                    "Linha": row_num,
                    "Tipo": "Data de protocolo de Ouvidoria inválida",
                    "Detalhe": clean_text(raw_protocol),
                })
                continue
            if protocol_date is None or protocol_date < base_start or protocol_date > end:
                continue
            raw_ouvidoria_protocol = row[protocolo_ouvidoria_col - 1].value if protocolo_ouvidoria_col <= len(row) else None
            raw_return = row[retorno_col - 1].value if retorno_col <= len(row) else None
            if not retorno_em_aberto(raw_return):
                continue
            records.append({
                "Empresa": source["empresa"],
                "Nº REC./REQ.": clean_text(row[source["rec_req"] - 1].value),
                "Tese": clean_text(row[source["tese"] - 1].value),
                "Responsável": clean_name(row[cols["responsavel"] - 1].value),
                "Parceiro": clean_text(row[0].value, "S/I"),
                "Município": clean_city(row[cols["municipio"] - 1].value),
                "Estado": clean_state(row[cols["municipio"]].value) if len(row) > cols["municipio"] else "S/I",
                "Data Protocolo Ouvidoria": protocol_date.isoformat(),
                "Protocolo de Ouvidoria": clean_text(raw_ouvidoria_protocol),
                "Data Retorno Ouvidoria": clean_text(raw_return, "S/I"),
                "Linha Original": row_num,
                "Arquivo": source["path"].name,
                "Coluna Protocolo Ouvidoria": protocolo_ouvidoria_col,
                "Coluna Retorno": retorno_col,
            })

    total = len(records)
    por_empresa = [
        {"Empresa": empresa, "Casos em Aberto": count, "% do Total": pct(count, total)}
        for empresa, count in Counter(r["Empresa"] for r in records).most_common()
    ]
    por_responsavel = [
        {"Responsável": resp, "Casos em Aberto": count, "% do Total": pct(count, total)}
        for resp, count in Counter(r["Responsável"] for r in records).most_common()
    ]
    por_municipio = [
        {"Parceiro": parceiro, "Município": municipio, "Estado": estado, "Casos em Aberto": count, "% do Total": pct(count, total)}
        for (parceiro, municipio, estado), count in Counter((r["Parceiro"], r["Município"], r["Estado"]) for r in records).most_common()
    ]
    return {
        "records": records,
        "por_empresa": por_empresa,
        "por_responsavel": por_responsavel,
        "por_municipio": por_municipio,
        "totais": {"Casos em Aberto": total, "Data Inicial Base": base_start.isoformat(), "Data Final Base": end.isoformat()},
        "inconsistencias": inconsistencias,
    }


def find_task_header(ws):
    for row_number, row_cells in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=False), start=1):
        normalized = [norm_header(c.value) for c in row_cells]
        if "DATA DE CONCLUSAO" in normalized and "PADRONIZACAO" in normalized and any(v in {"QTD", "QUANTIDADE"} for v in normalized):
            cols = {}
            for idx, value in enumerate(normalized, start=1):
                if value == "PADRONIZACAO":
                    cols["padronizacao"] = idx
                elif value in {"QTD", "QUANTIDADE"}:
                    cols["quantidade"] = idx
                    cols["quantidade_nome"] = value
                elif value == "DATA DE CONCLUSAO":
                    cols["data_conclusao"] = idx
            return row_number, cols
    return None, None


def web_post_json(path, payload, cookie=None, timeout=45):
    raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    headers = {"Content-Type": "application/json; charset=utf-8"}
    if cookie:
        headers["Cookie"] = cookie
    request = urllib.request.Request(WEB_TASKS_URL + path, data=raw, headers=headers, method="POST")
    with urllib.request.urlopen(request, timeout=timeout) as response:
        data = json.loads(response.read().decode("utf-8"))
        if "error" in data:
            raise ValueError(data["error"])
        return data, response.headers


def fetch_web_task_records(start, end):
    username = os.environ.get("TASK_WEB_USER", "").strip()
    password = os.environ.get("TASK_WEB_PASSWORD", "")
    if not username or not password:
        raise ValueError("Login/senha do sistema web não informados.")

    _, headers = web_post_json("/api/login", {"username": username, "password": password})
    set_cookie = headers.get("Set-Cookie", "")
    cookie = set_cookie.split(";", 1)[0]
    if not cookie:
        raise ValueError("Login realizado, mas cookie de sessão não foi retornado.")

    report, _ = web_post_json(
        "/api/reports/period",
        {"start": start.isoformat(), "end": end.isoformat(), "assignee_id": ""},
        cookie=cookie,
    )
    details = report.get("completed", {}).get("details", [])
    records = []
    for idx, task in enumerate(details, start=1):
        raw_date = task.get("completed_at")
        date_value = parse_date(raw_date)
        if date_value is None or date_value < start or date_value > end:
            continue
        records.append({
            "Analista": clean_name(task.get("assignee_name")),
            "Padronização": clean_text(task.get("standard"), "SEM PADRONIZAÇÃO") or "SEM PADRONIZAÇÃO",
            "Quantidade": parse_number(task.get("quantity")) or 0.0,
            "Data de Conclusão": date_value.isoformat(),
            "Linha Original": task.get("id") or idx,
            "Arquivo": "controle-equipe.onrender.com",
            "Aba": "API /api/reports/period",
        })
    return records, {
        "completed_total_api": report.get("completed", {}).get("total", len(details)),
        "completed_details_api": len(details),
        "started_total_api": report.get("started", {}).get("total", 0),
    }


def summarize_tarefas(records, inconsistencias):
    totals = defaultdict(float)
    line_counts = defaultdict(int)
    ranking_map = defaultdict(lambda: defaultdict(float))
    for record in records:
        analista = record.get("Analista") or "S/I"
        quantidade = record.get("Quantidade") or 0.0
        padronizacao = (
            record.get("Padronização")
            or record.get("Padroniza??o")
            or record.get("PadronizaÃ§Ã£o")
            or "SEM PADRONIZAÇÃO"
        )
        data_conclusao = (
            record.get("Data de Conclusão")
            or record.get("Data de Conclus?o")
            or record.get("Data de ConclusÃ£o")
            or ""
        )
        record["Analista"] = analista
        record["Quantidade"] = quantidade
        record["Padronização"] = padronizacao
        record["Data de Conclusão"] = data_conclusao
        totals[analista] += quantidade
        line_counts[analista] += 1
        ranking_map[analista][padronizacao] += quantidade
    total = sum(totals.values())
    resumo = []
    ranking = []
    analistas = ["Anderson", "Débora", "Gonzaga", "Nickolas"]
    for analista in sorted(set(analistas) | set(totals.keys())):
        resumo.append({
            "Analista": analista,
            "Total de Atividades": totals.get(analista, 0.0),
            "Linhas Consideradas": line_counts.get(analista, 0),
            "% do Total": pct(totals.get(analista, 0.0), total),
        })
        for pos, (atividade, qty) in enumerate(sorted(ranking_map[analista].items(), key=lambda kv: (-kv[1], kv[0])), start=1):
            ranking.append({
                "Analista": analista,
                "Ranking": pos,
                "Atividade": atividade,
                "Quantidade": qty,
                "Top 6": "Sim" if pos <= 6 else "Não",
            })
    return {
        "records": records,
        "resumo": resumo,
        "ranking": ranking,
        "total": total,
        "inconsistencias": inconsistencias,
    }



def apurar_tarefas_planilhas(start, end, inconsistencias):
    records = []
    for path in TAREFAS_FILES:
        analista = analista_from_file(path)
        if not path.exists():
            inconsistencias.append({
                "Origem": "Atividades",
                "Arquivo": path.name,
                "Aba": "",
                "Linha": "",
                "Tipo": "Arquivo de tarefas n?o localizado",
                "Detalhe": str(path),
            })
            continue
        wb = load_workbook_safe(path, read_only=True, data_only=True)
        found = False
        for ws in wb.worksheets:
            header_row, cols = find_task_header(ws)
            if not cols:
                continue
            found = True
            inconsistencias.append({
                "Origem": "Atividades",
                "Arquivo": path.name,
                "Aba": ws.title,
                "Linha": header_row,
                "Tipo": "Coluna de quantidade identificada",
                "Detalhe": cols["quantidade_nome"],
            })
            for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=False), start=header_row + 1):
                raw_date = row[cols["data_conclusao"] - 1].value
                date_value = parse_date(raw_date)
                if should_report_invalid_date(raw_date, include_question=True) and date_value is None:
                    inconsistencias.append({
                        "Origem": "Atividades",
                        "Arquivo": path.name,
                        "Aba": ws.title,
                        "Linha": row_num,
                        "Tipo": "Data de conclus?o fora do padr?o",
                        "Detalhe": clean_text(raw_date),
                    })
                    continue
                if date_value is None or date_value < start or date_value > end:
                    continue
                pad = clean_text(row[cols["padronizacao"] - 1].value)
                qty = parse_number(row[cols["quantidade"] - 1].value)
                if not pad:
                    inconsistencias.append({
                        "Origem": "Atividades",
                        "Arquivo": path.name,
                        "Aba": ws.title,
                        "Linha": row_num,
                        "Tipo": "Linha no per?odo sem PADRONIZA??O",
                        "Detalhe": "",
                    })
                    pad = "SEM PADRONIZA??O"
                if qty is None:
                    inconsistencias.append({
                        "Origem": "Atividades",
                        "Arquivo": path.name,
                        "Aba": ws.title,
                        "Linha": row_num,
                        "Tipo": "Linha no per?odo sem QTD/QUANTIDADE v?lida",
                        "Detalhe": clean_text(row[cols["quantidade"] - 1].value),
                    })
                    qty = 0.0
                records.append({
                    "Analista": analista,
                    "Padroniza??o": pad,
                    "Quantidade": qty,
                    "Data de Conclus?o": date_value.isoformat(),
                    "Linha Original": row_num,
                    "Arquivo": path.name,
                    "Aba": ws.title,
                })
        if not found:
            inconsistencias.append({
                "Origem": "Atividades",
                "Arquivo": path.name,
                "Aba": "",
                "Linha": "",
                "Tipo": "Aba com colunas obrigat?rias n?o encontrada",
                "Detalhe": "PADRONIZA??O, DATA DE CONCLUS?O e QTD/QUANTIDADE",
            })
    return records


def apurar_tarefas(start, end):
    records = []
    inconsistencias = []

    if start < TAREFAS_WEB_START:
        sheet_start = start
        sheet_end = min(end, date(2026, 6, 7))
        if sheet_start <= sheet_end:
            sheet_records = apurar_tarefas_planilhas(sheet_start, sheet_end, inconsistencias)
            records.extend(sheet_records)
            inconsistencias.append({
                "Origem": "Atividades",
                "Arquivo": "Planilhas de tarefas",
                "Aba": "",
                "Linha": "",
                "Tipo": "Fonte de atividades",
                "Detalhe": f"Planilhas usadas de {sheet_start.strftime('%d/%m/%Y')} a {sheet_end.strftime('%d/%m/%Y')}; consideradas: {len(sheet_records)}.",
            })

    if end >= TAREFAS_WEB_START:
        web_start = max(start, TAREFAS_WEB_START)
        web_end = end
        try:
            web_records, web_meta = fetch_web_task_records(web_start, web_end)
            records.extend(web_records)
            inconsistencias.append({
                "Origem": "Atividades",
                "Arquivo": WEB_TASKS_URL,
                "Aba": "API",
                "Linha": "",
                "Tipo": "Fonte de atividades",
                "Detalhe": f"Sistema web usado de {web_start.strftime('%d/%m/%Y')} a {web_end.strftime('%d/%m/%Y')}. Conclu?das API: {web_meta['completed_total_api']}; detalhes: {web_meta['completed_details_api']}; iniciadas API: {web_meta['started_total_api']}; consideradas: {len(web_records)}.",
            })
        except Exception as exc:
            inconsistencias.append({
                "Origem": "Atividades",
                "Arquivo": WEB_TASKS_URL,
                "Aba": "API",
                "Linha": "",
                "Tipo": "Falha ao importar atividades do sistema web",
                "Detalhe": str(exc),
            })

    return summarize_tarefas(records, inconsistencias)

def empty_nf_result(inconsistencias=None):
    return {
        "records": [],
        "anderson": [],
        "por_responsavel": [],
        "por_empresa_anderson": [
            {"Empresa": company, "Quantidade": 0, "Valor Total": 0.0, "Valor dos Processos": 0.0}
            for company in ["INOVVE", "HLA", "GRID", "FDOJ"]
        ],
        "totais": {
            "Quantidade Total": 0,
            "Valor Total": 0.0,
            "Valor dos Processos": 0.0,
            "Quantidade Anderson": 0,
            "Valor Anderson": 0.0,
            "Valor Processos Anderson": 0.0,
        },
        "inconsistencias": inconsistencias or [],
    }


def apurar_nf(start, end):
    if not NF_FILE.exists():
        return empty_nf_result([{
            "Origem": "Faturamento NF",
            "Arquivo": NF_FILE.name,
            "Aba": "NFS",
            "Linha": "",
            "Tipo": "Arquivo não encontrado",
            "Detalhe": str(NF_FILE),
        }])
    wb = load_workbook_safe(NF_FILE, read_only=True, data_only=True)
    if "NFS" not in wb.sheetnames:
        return empty_nf_result([{
            "Origem": "Faturamento NF",
            "Arquivo": NF_FILE.name,
            "Aba": "NFS",
            "Linha": "",
            "Tipo": "Aba não encontrada",
            "Detalhe": f"Abas disponíveis: {', '.join(wb.sheetnames)}",
        }])
    ws = wb["NFS"]
    cols = {
        "data": 3,
        "responsavel": 4,
        "municipio": 5,
        "estado": 6,
        "contrato": 7,
        "valor_processo": 11,
        "aliquota": 15,
        "valor_servico": 16,
    }
    records = []
    inconsistencias = []
    for row_num, row in enumerate(ws.iter_rows(min_row=5, values_only=False), start=5):
        date_value = parse_date(row[cols["data"] - 1].value)
        if date_value is None or date_value < start or date_value > end:
            continue
        valor_processo = parse_money(row[cols["valor_processo"] - 1].value)
        aliquota = parse_rate(row[cols["aliquota"] - 1].value)
        valor_servico = parse_money(row[cols["valor_servico"] - 1].value)
        calc = valor_processo * aliquota if aliquota is not None else None
        diff = None if calc is None else round(valor_servico - calc, 2)
        if calc is not None and abs(diff) > 0.05:
            inconsistencias.append({
                "Origem": "Faturamento NF",
                "Arquivo": NF_FILE.name,
                "Aba": "NFS",
                "Linha": row_num,
                "Tipo": "SERVIÇO (R$) difere de VALOR x SERVIÇO (%)",
                "Detalhe": f"Serviço={valor_servico:.2f}; calculado={calc:.2f}; diferença={diff:.2f}",
            })
        records.append({
            "Data de Envio": date_value.isoformat(),
            "Responsável": clean_name(row[cols["responsavel"] - 1].value),
            "Município": clean_city(row[cols["municipio"] - 1].value),
            "Estado": clean_state(row[cols["estado"] - 1].value),
            "Empresa": company_from_contract(row[cols["contrato"] - 1].value),
            "Contrato": clean_text(row[cols["contrato"] - 1].value),
            "Valor do Processo": valor_processo,
            "Alíquota do Serviço": aliquota,
            "Valor do Retorno/Serviço": valor_servico,
            "Linha Original": row_num,
        })
    total_servico = sum(r["Valor do Retorno/Serviço"] for r in records)
    total_processo = sum(r["Valor do Processo"] for r in records)
    anderson = [r for r in records if r["Responsável"] == "Anderson"]
    by_resp_acc = defaultdict(lambda: {"Quantidade": 0, "Valor Total": 0.0, "Valor dos Processos": 0.0})
    for record in records:
        acc = by_resp_acc[record["Responsável"]]
        acc["Quantidade"] += 1
        acc["Valor Total"] += record["Valor do Retorno/Serviço"]
        acc["Valor dos Processos"] += record["Valor do Processo"]
    by_resp = [
        {
            "Responsável": resp,
            "Quantidade": vals["Quantidade"],
            "Valor Total": vals["Valor Total"],
            "% do Valor Total": pct(vals["Valor Total"], total_servico),
            "Valor dos Processos": vals["Valor dos Processos"],
        }
        for resp, vals in sorted(by_resp_acc.items(), key=lambda kv: (-kv[1]["Valor Total"], kv[0]))
    ]
    by_company_acc = defaultdict(lambda: {"Quantidade": 0, "Valor Total": 0.0, "Valor dos Processos": 0.0})
    for record in anderson:
        if record["Empresa"] in {"INOVVE", "HLA", "GRID", "FDOJ"}:
            acc = by_company_acc[record["Empresa"]]
            acc["Quantidade"] += 1
            acc["Valor Total"] += record["Valor do Retorno/Serviço"]
            acc["Valor dos Processos"] += record["Valor do Processo"]
    by_company = [
        {
            "Empresa": company,
            "Quantidade": by_company_acc[company]["Quantidade"],
            "Valor Total": by_company_acc[company]["Valor Total"],
            "Valor dos Processos": by_company_acc[company]["Valor dos Processos"],
        }
        for company in ["INOVVE", "HLA", "GRID", "FDOJ"]
    ]
    return {
        "records": records,
        "anderson": anderson,
        "por_responsavel": by_resp,
        "por_empresa_anderson": by_company,
        "totais": {
            "Quantidade Total": len(records),
            "Valor Total": total_servico,
            "Valor dos Processos": total_processo,
            "Quantidade Anderson": len(anderson),
            "Valor Anderson": sum(r["Valor do Retorno/Serviço"] for r in anderson),
            "Valor Processos Anderson": sum(r["Valor do Processo"] for r in anderson),
        },
        "inconsistencias": inconsistencias,
    }


def apurar_nf_historico(year, month):
    wb = load_workbook_safe(NF_FILE, read_only=True, data_only=True)
    ws = wb["NFS"]
    cols = {
        "data": 3,
        "responsavel": 4,
        "valor_processo": 11,
        "valor_servico": 16,
    }
    acc = defaultdict(lambda: {"Quantidade": 0, "Valor dos Processos": 0.0, "Valor Serviço": 0.0})
    for row in ws.iter_rows(min_row=5, values_only=False):
        date_value = parse_date(row[cols["data"] - 1].value)
        if date_value is None:
            continue
        if clean_name(row[cols["responsavel"] - 1].value) != "Anderson":
            continue
        if date_value.year != 2025 and not (date_value.year == year and date_value.month <= month):
            continue
        key = f"{date_value.year}-{date_value.month:02d}"
        acc[key]["Quantidade"] += 1
        acc[key]["Valor dos Processos"] += parse_money(row[cols["valor_processo"] - 1].value)
        acc[key]["Valor Serviço"] += parse_money(row[cols["valor_servico"] - 1].value)

    meses_2025 = []
    for m in range(1, 13):
        key = f"2025-{m:02d}"
        meses_2025.append({
            "Ano": 2025,
            "Mês": m,
            "Período": key,
            "Quantidade": acc[key]["Quantidade"],
            "Valor dos Processos": acc[key]["Valor dos Processos"],
            "Valor Serviço": acc[key]["Valor Serviço"],
        })
    media_2025 = sum(r["Valor dos Processos"] for r in meses_2025) / 12

    ano_atual = []
    for m in range(1, month + 1):
        key = f"{year}-{m:02d}"
        ano_atual.append({
            "Ano": year,
            "Mês": m,
            "Período": key,
            "Quantidade": acc[key]["Quantidade"],
            "Valor dos Processos": acc[key]["Valor dos Processos"],
            "Valor Serviço": acc[key]["Valor Serviço"],
        })

    values = [r["Valor dos Processos"] for r in ano_atual]
    n = len(values)
    if n >= 2:
        xs = list(range(1, n + 1))
        avg_x = sum(xs) / n
        avg_y = sum(values) / n
        denom = sum((x - avg_x) ** 2 for x in xs)
        slope = sum((x - avg_x) * (y - avg_y) for x, y in zip(xs, values)) / denom if denom else 0
        intercept = avg_y - slope * avg_x
    else:
        slope = 0
        intercept = values[0] if values else 0
    for idx, row in enumerate(ano_atual, start=1):
        row["Tendência Linear"] = intercept + slope * idx
        row["Média Mensal 2025"] = media_2025

    return {
        "media_mensal_2025_processos": media_2025,
        "tendencia": {"inclinação": slope, "intercepto": intercept},
        "meses_2025": meses_2025,
        "ano_atual": ano_atual,
    }


def apurar_deferimentos(start, end):
    records = []
    inconsistencias = []
    for src in DEFERIMENTOS_SOURCES:
        if not src["path"].exists():
            inconsistencias.append({
                "Origem": "Deferimentos",
                "Arquivo": src["path"].name,
                "Aba": src["sheet"],
                "Linha": "",
                "Tipo": "Arquivo não encontrado",
                "Detalhe": str(src["path"]),
            })
            continue
        wb = load_workbook_safe(src["path"], read_only=True, data_only=True, keep_vba=True)
        if src["sheet"] not in wb.sheetnames:
            inconsistencias.append({
                "Origem": "Deferimentos",
                "Arquivo": src["path"].name,
                "Aba": src["sheet"],
                "Linha": "",
                "Tipo": "Aba não encontrada",
                "Detalhe": f"Abas disponíveis: {', '.join(wb.sheetnames)}",
            })
            continue
        ws = wb[src["sheet"]]
        detected_header_row, detected_cols = find_deferimento_payment_columns(ws, src["header_row"])
        date_col = detected_cols.get("data_deferimento") or src["data_deferimento"]
        value_col = detected_cols.get("valor_deferido") or src["valor_deferido"]
        for row_num, row in enumerate(ws.iter_rows(min_row=detected_header_row + 1, values_only=False), start=detected_header_row + 1):
            municipio = clean_city(row[src["municipio"] - 1].value)
            estado = clean_state(row[src["municipio"]].value) if len(row) > src["municipio"] else "S/I"
            if not allowed_city(src, municipio):
                continue
            raw_date = row[date_col - 1].value
            date_value = parse_date(raw_date)
            raw_value = row[value_col - 1].value
            valor = parse_money(raw_value)
            if should_report_invalid_date(raw_date) and date_value is None:
                inconsistencias.append({
                    "Origem": "Deferimentos",
                    "Arquivo": src["path"].name,
                    "Aba": src["sheet"],
                    "Linha": row_num,
                    "Tipo": "Data de deferimento fora do padrão",
                    "Detalhe": clean_text(raw_date),
                })
            if date_value is None or date_value < start or date_value > end:
                continue
            if raw_value in (None, ""):
                inconsistencias.append({
                    "Origem": "Deferimentos",
                    "Arquivo": src["path"].name,
                    "Aba": src["sheet"],
                    "Linha": row_num,
                    "Tipo": "Deferimento no período sem valor deferido",
                    "Detalhe": "",
                })
            records.append({
                "Empresa": src["empresa"],
                "Parceiro": clean_text(row[0].value, "S/I"),
                "Município": municipio,
                "Estado": estado,
                "Data Deferimento": date_value.isoformat(),
                "Valor Deferido": valor,
                "Reclamação": clean_text(row[src["reclamacao"] - 1].value),
                "Tese": clean_text(row[src["tese"] - 1].value),
                "Linha Original": row_num,
                "Arquivo": src["path"].name,
            })
    total = sum(r["Valor Deferido"] for r in records)
    by_company_acc = defaultdict(lambda: {"Quantidade": 0, "Valor Deferido": 0.0})
    by_city_acc = defaultdict(lambda: {"Quantidade": 0, "Valor Deferido": 0.0})
    for record in records:
        by_company_acc[record["Empresa"]]["Quantidade"] += 1
        by_company_acc[record["Empresa"]]["Valor Deferido"] += record["Valor Deferido"]
        key = (record["Empresa"], record["Parceiro"], record["Município"], record["Estado"])
        by_city_acc[key]["Quantidade"] += 1
        by_city_acc[key]["Valor Deferido"] += record["Valor Deferido"]
    by_company = [
        {
            "Empresa": company,
            "Quantidade": by_company_acc[company]["Quantidade"],
            "Valor Deferido": by_company_acc[company]["Valor Deferido"],
            "% do Valor": pct(by_company_acc[company]["Valor Deferido"], total),
        }
        for company in EMPRESAS_RECLAMACOES
    ]
    by_city = [
        {"Empresa": key[0], "Parceiro": key[1], "Município": key[2], "Estado": key[3], "Quantidade": vals["Quantidade"], "Valor Deferido": vals["Valor Deferido"]}
        for key, vals in sorted(by_city_acc.items(), key=lambda kv: (-kv[1]["Valor Deferido"], kv[0][0], kv[0][1], kv[0][2], kv[0][3]))
    ]
    return {
        "records": sorted(records, key=lambda r: (r["Empresa"], r["Parceiro"], r["Município"], r["Estado"], r["Data Deferimento"], r["Linha Original"])),
        "por_empresa": by_company,
        "por_municipio": by_city,
        "totais": {"Quantidade": len(records), "Valor Deferido": total},
        "inconsistencias": inconsistencias,
    }


def find_deferimento_payment_columns(ws, preferred_header_row):
    aliases = {
        "data_deferimento": ("DATA", "DEFERIMENTO"),
        "valor_deferido": ("VALOR", "DEFERIDO"),
        "valor_pago": ("VALOR", "PAGO"),
    }
    search_rows = list(range(max(1, preferred_header_row - 2), min(ws.max_row, preferred_header_row + 2) + 1))
    for row_num in search_rows:
        headers = {col: norm_header(ws.cell(row_num, col).value) for col in range(1, min(ws.max_column, 40) + 1)}
        found = {}
        for key, terms in aliases.items():
            found[key] = next((col for col, header in headers.items() if all(term in header for term in terms)), None)
        if all(found.values()):
            return row_num, found
    return preferred_header_row, {
        "data_deferimento": None,
        "valor_deferido": None,
        "valor_pago": None,
    }


def apurar_deferimentos_pendentes(end):
    start = date(2025, 1, 1)
    records = []
    inconsistencias = []
    for src in DEFERIMENTOS_SOURCES:
        if not src["path"].exists():
            inconsistencias.append({
                "Origem": "Deferimentos pendentes",
                "Arquivo": src["path"].name,
                "Aba": src["sheet"],
                "Linha": "",
                "Tipo": "Arquivo não encontrado",
                "Detalhe": str(src["path"]),
            })
            continue
        wb = load_workbook_safe(src["path"], read_only=True, data_only=True, keep_vba=True)
        if src["sheet"] not in wb.sheetnames:
            inconsistencias.append({
                "Origem": "Deferimentos pendentes",
                "Arquivo": src["path"].name,
                "Aba": src["sheet"],
                "Linha": "",
                "Tipo": "Aba não encontrada",
                "Detalhe": f"Abas disponíveis: {', '.join(wb.sheetnames)}",
            })
            continue
        ws = wb[src["sheet"]]
        header_row, cols = find_deferimento_payment_columns(ws, src["header_row"])
        if not all(cols.values()):
            inconsistencias.append({
                "Origem": "Deferimentos pendentes",
                "Arquivo": src["path"].name,
                "Aba": src["sheet"],
                "Linha": header_row,
                "Tipo": "Colunas de deferimento/pagamento não localizadas",
                "Detalhe": "Esperado: DATA DEFERIMENTO, VALOR DEFERIDO e VALOR PAGO ATÉ AGORA",
            })
            continue
        for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=False), start=header_row + 1):
            municipio = clean_city(row[src["municipio"] - 1].value)
            estado = clean_state(row[src["municipio"]].value) if len(row) > src["municipio"] else "S/I"
            if not allowed_city(src, municipio):
                continue
            date_value = parse_date(row[cols["data_deferimento"] - 1].value)
            if date_value is None or date_value < start or date_value > end:
                continue
            valor_deferido = parse_money(row[cols["valor_deferido"] - 1].value)
            valor_pago = parse_money(row[cols["valor_pago"] - 1].value)
            saldo = max(valor_deferido - valor_pago, 0.0)
            if valor_deferido <= 0 or saldo < 0.01:
                continue
            situacao = "Não pago" if valor_pago < 0.01 else "Pago parcialmente"
            records.append({
                "Empresa": src["empresa"],
                "Parceiro": clean_text(row[0].value, "S/I"),
                "Município": municipio,
                "Estado": estado,
                "Data Deferimento": date_value.isoformat(),
                "Reclamação": clean_text(row[src["reclamacao"] - 1].value),
                "Tese": clean_text(row[src["tese"] - 1].value),
                "Situação do Pagamento": situacao,
                "Valor Deferido": valor_deferido,
                "Valor Pago Até Agora": valor_pago,
                "Saldo Pendente": saldo,
                "% Pago": valor_pago / valor_deferido if valor_deferido else 0,
                "Linha Original": row_num,
                "Arquivo": src["path"].name,
            })

    def summarize(key_name):
        acc = defaultdict(lambda: {"Quantidade": 0, "Valor Deferido": 0.0, "Valor Pago Até Agora": 0.0, "Saldo Pendente": 0.0})
        for record in records:
            item = acc[record[key_name]]
            item["Quantidade"] += 1
            item["Valor Deferido"] += record["Valor Deferido"]
            item["Valor Pago Até Agora"] += record["Valor Pago Até Agora"]
            item["Saldo Pendente"] += record["Saldo Pendente"]
        return [
            {key_name: key, **values}
            for key, values in sorted(acc.items(), key=lambda item: (-item[1]["Saldo Pendente"], item[0]))
        ]

    total_deferido = sum(r["Valor Deferido"] for r in records)
    total_pago = sum(r["Valor Pago Até Agora"] for r in records)
    total_saldo = sum(r["Saldo Pendente"] for r in records)
    return {
        "records": sorted(records, key=lambda r: (-r["Saldo Pendente"], r["Empresa"], r["Parceiro"], r["Município"], r["Estado"])),
        "por_situacao": summarize("Situação do Pagamento"),
        "por_empresa": summarize("Empresa"),
        "totais": {
            "Quantidade": len(records),
            "Valor Deferido": total_deferido,
            "Valor Pago Até Agora": total_pago,
            "Saldo Pendente": total_saldo,
        },
        "inconsistencias": inconsistencias,
    }


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="1F4E78")
SUBTITLE_FONT = Font(bold=True, color="444444")
THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_title(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUBTITLE_FONT


def write_table(ws, start_row, start_col, headers, rows, table_name):
    for offset, header in enumerate(headers):
        cell = ws.cell(start_row, start_col + offset, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for row_offset, row in enumerate(rows, start=1):
        for col_offset, header in enumerate(headers):
            cell = ws.cell(start_row + row_offset, start_col + col_offset, row.get(header, ""))
            cell.border = BORDER
            if header.startswith("%") or header in {"% do Total", "% do Valor", "% do Valor Total", "Alíquota do Serviço"}:
                cell.number_format = "0.00%"
            elif header in {
                "Valor",
                "Valor Total",
                "Valor dos Processos",
                "Valor Serviço",
                "Valor Anderson",
                "Valor Processos Anderson",
                "Valor do Processo",
                "Valor do Retorno/Serviço",
                "Valor Deferido",
                "Valor Pago Até Agora",
                "Saldo Pendente",
                "Média Mensal 2025",
                "Tendência Linear",
            }:
                cell.number_format = '"R$" #,##0.00'
            elif header in {"Atual", "Mês Anterior", "Variação"} and row.get("Unidade") == "Valor":
                cell.number_format = '"R$" #,##0.00'
            elif header in {"Total de Atividades", "Quantidade"}:
                cell.number_format = "#,##0.00"
    end_row = start_row + max(len(rows), 1)
    end_col = start_col + len(headers) - 1
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    return end_row + 2


def auto_width(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 46)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"


def build_workbook(month, year, start, end, ingressos, tarefas, nf, deferimentos, inconsistencias, comparativo=None, prev_start=None, prev_end=None, nf_historico=None, ouvidoria_abertos=None, deferimentos_pendentes=None, output_stem=None, title="Apuração mensal consolidada"):
    wb = Workbook()
    wb.remove(wb.active)
    subtitle = f"Período: {start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')}"

    ws = wb.create_sheet("Resumo Geral")
    write_title(ws, title, subtitle)
    rows = [
        {"Apuração": "Ingressos de Reclamações", "Indicador": "Total", "Quantidade": ingressos["totais"]["Total"], "Valor": ""},
        {"Apuração": "Ingressos de Reclamações", "Indicador": "Concessionária", "Quantidade": ingressos["totais"]["Concessionária"], "Valor": ""},
        {"Apuração": "Ingressos de Reclamações", "Indicador": "Ouvidoria", "Quantidade": ingressos["totais"]["Ouvidoria"], "Valor": ""},
        {"Apuração": "Ingressos de Reclamações", "Indicador": "Ouvidoria ANEEL", "Quantidade": ingressos["totais"]["Ouvidoria ANEEL"], "Valor": ""},
        {"Apuração": "Ingressos de Reclamações", "Indicador": "Processo ADM", "Quantidade": ingressos["totais"]["Processo ADM"], "Valor": ""},
        {"Apuração": "Atividades Gerais", "Indicador": "Total de atividades", "Quantidade": tarefas["total"], "Valor": ""},
        {"Apuração": "Faturamento NF", "Indicador": "Registros Anderson", "Quantidade": nf["totais"]["Quantidade Anderson"], "Valor": nf["totais"]["Valor Anderson"]},
        {"Apuração": "Faturamento NF", "Indicador": "Valor dos processos Anderson", "Quantidade": "", "Valor": nf["totais"]["Valor Processos Anderson"]},
        {"Apuração": "Deferimentos", "Indicador": "Quantidade", "Quantidade": deferimentos["totais"]["Quantidade"], "Valor": deferimentos["totais"]["Valor Deferido"]},
    ]
    if ouvidoria_abertos is not None:
        rows.append({
            "Apuração": "Ouvidoria em aberto",
            "Indicador": "Protocolos sem retorno desde 01/01/2025",
            "Quantidade": ouvidoria_abertos["totais"]["Casos em Aberto"],
            "Valor": "",
        })
    if deferimentos_pendentes is not None:
        rows.append({
            "Apuração": "Deferimentos pendentes",
            "Indicador": "Saldo não pago desde 01/01/2025",
            "Quantidade": deferimentos_pendentes["totais"]["Quantidade"],
            "Valor": deferimentos_pendentes["totais"]["Saldo Pendente"],
        })
    write_table(ws, 4, 1, ["Apuração", "Indicador", "Quantidade", "Valor"], rows, "ResumoGeral")
    auto_width(ws)

    if comparativo is not None:
        ws = wb.create_sheet("Comparativo Mensal")
        prev_sub = f"Comparação: {start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')} vs {prev_start.strftime('%d/%m/%Y')} a {prev_end.strftime('%d/%m/%Y')}"
        write_title(ws, "Comparativo com mês anterior", prev_sub)
        write_table(ws, 4, 1, ["Grupo", "Indicador", "Atual", "Mês Anterior", "Variação", "% Variação", "Status", "Unidade"], comparativo, "ComparativoMensal")
        auto_width(ws)

        comparison_sheets = [
            ("Comp Ingressos", "Comparativo - ingressos", ["Ingressos", "Ingressos por empresa", "Ingressos por colaborador"]),
            ("Comp Atividades", "Comparativo - atividades gerais", ["Atividades", "Atividades por analista"]),
            ("Comp NF", "Comparativo - processos enviados para NF", ["Faturamento NF", "NF Anderson por empresa"]),
            ("Comp Deferimentos", "Comparativo - deferimentos", ["Deferimentos", "Deferimentos por empresa"]),
        ]
        for sheet_name, title, groups in comparison_sheets:
            ws = wb.create_sheet(sheet_name)
            write_title(ws, title, prev_sub)
            rows_cmp = [row for row in comparativo if row["Grupo"] in groups]
            write_table(ws, 4, 1, ["Grupo", "Indicador", "Atual", "Mês Anterior", "Variação", "% Variação", "Status", "Unidade"], rows_cmp, sheet_name.replace(" ", ""))
            auto_width(ws)

    ws = wb.create_sheet("Ingressos Reclamações")
    write_title(ws, "Ingressos de reclamações", subtitle)
    next_row = write_table(ws, 4, 1, ["Empresa", "Concessionária", "Ouvidoria", "Ouvidoria ANEEL", "Processo ADM", "Total", "% do Total"], ingressos["resumo_empresa"], "IngressosEmpresa")
    next_row = write_table(ws, next_row, 1, ["Responsável", "Total", "% do Total"], ingressos["responsaveis"], "IngressosResponsavel")
    write_table(ws, next_row, 1, ["Parceiro", "Município", "Estado", "Total", "% do Total"], ingressos["municipios"], "IngressosMunicipio")
    auto_width(ws)

    ws = wb.create_sheet("Ingressos Base Detalhada")
    write_title(ws, "Base detalhada - ingressos", subtitle)
    write_table(ws, 4, 1, ["Empresa", "Tipo de Ingresso", "Nº REC./REQ.", "Tese", "Responsável", "Parceiro", "Município", "Estado", "Data", "Linha Original", "Arquivo"], ingressos["records"], "IngressosBase")
    auto_width(ws)

    if ouvidoria_abertos is not None:
        ws = wb.create_sheet("Ouvidoria Em Aberto")
        write_title(ws, "Ouvidoria em aberto", f"Base: protocolos de 01/01/2025 a {end.strftime('%d/%m/%Y')} sem data de retorno")
        next_row = write_table(ws, 4, 1, ["Empresa", "Casos em Aberto", "% do Total"], ouvidoria_abertos["por_empresa"], "OuvAbertoEmpresa")
        next_row = write_table(ws, next_row, 1, ["Responsável", "Casos em Aberto", "% do Total"], ouvidoria_abertos["por_responsavel"], "OuvAbertoResp")
        next_row = write_table(ws, next_row, 1, ["Parceiro", "Município", "Estado", "Casos em Aberto", "% do Total"], ouvidoria_abertos["por_municipio"], "OuvAbertoMun")
        write_table(ws, next_row, 1, ["Empresa", "Nº REC./REQ.", "Tese", "Responsável", "Parceiro", "Município", "Estado", "Data Protocolo Ouvidoria", "Protocolo de Ouvidoria", "Data Retorno Ouvidoria", "Linha Original", "Arquivo", "Coluna Protocolo Ouvidoria", "Coluna Retorno"], ouvidoria_abertos["records"], "OuvAbertoBase")
        auto_width(ws)

    ws = wb.create_sheet("Atividades Gerais")
    write_title(ws, "Atividades gerais", subtitle)
    write_table(ws, 4, 1, ["Analista", "Total de Atividades", "Linhas Consideradas", "% do Total"], tarefas["resumo"], "AtividadesResumo")
    auto_width(ws)

    ws = wb.create_sheet("Ranking Atividades")
    write_title(ws, "Ranking de atividades por analista", subtitle)
    write_table(ws, 4, 1, ["Analista", "Ranking", "Atividade", "Quantidade", "Top 6"], tarefas["ranking"], "AtividadesRanking")
    auto_width(ws)

    ws = wb.create_sheet("Faturamento NF")
    write_title(ws, "Faturamento NF", subtitle)
    nf_summary = [
        {"Indicador": "Quantidade total no período", "Valor": nf["totais"]["Quantidade Total"]},
        {"Indicador": "Valor total enviado no período", "Valor": nf["totais"]["Valor Total"]},
        {"Indicador": "Quantidade enviada por Anderson", "Valor": nf["totais"]["Quantidade Anderson"]},
        {"Indicador": "Valor enviado por Anderson", "Valor": nf["totais"]["Valor Anderson"]},
        {"Indicador": "Valor dos processos de Anderson", "Valor": nf["totais"]["Valor Processos Anderson"]},
    ]
    next_row = write_table(ws, 4, 1, ["Indicador", "Valor"], nf_summary, "NFSResumo")
    next_row = write_table(ws, next_row, 1, ["Responsável", "Quantidade", "Valor Total", "% do Valor Total", "Valor dos Processos"], nf["por_responsavel"], "NFResponsavel")
    next_row = write_table(ws, next_row, 1, ["Empresa", "Quantidade", "Valor Total", "Valor dos Processos"], nf["por_empresa_anderson"], "NFEmpresaAnderson")
    write_table(ws, next_row, 1, ["Data de Envio", "Responsável", "Município", "Estado", "Empresa", "Contrato", "Valor do Processo", "Alíquota do Serviço", "Valor do Retorno/Serviço", "Linha Original"], nf["anderson"], "NFDetalheAnderson")
    auto_width(ws)

    if nf_historico is not None:
        ws = wb.create_sheet("NF Histórico")
        write_title(ws, "Histórico NF - Anderson", "Valor dos processos enviados para geração de NF")
        resumo_hist = [
            {"Indicador": "Média mensal 2025", "Valor": nf_historico["media_mensal_2025_processos"]},
            {"Indicador": "Meses 2025 considerados", "Valor": 12},
            {"Indicador": "Meses do ano atual considerados", "Valor": len(nf_historico["ano_atual"])},
        ]
        next_row = write_table(ws, 4, 1, ["Indicador", "Valor"], resumo_hist, "NFHistResumo")
        rows_2025 = [
            {
                "Ano": r["Ano"],
                "Mês": r["Mês"],
                "Período": r["Período"],
                "Quantidade": r["Quantidade"],
                "Valor dos Processos": r["Valor dos Processos"],
                "Valor Serviço": r["Valor Serviço"],
            }
            for r in nf_historico["meses_2025"]
        ]
        next_row = write_table(ws, next_row, 1, ["Ano", "Mês", "Período", "Quantidade", "Valor dos Processos", "Valor Serviço"], rows_2025, "NFHist2025")
        rows_current = [
            {
                "Ano": r["Ano"],
                "Mês": r["Mês"],
                "Período": r["Período"],
                "Quantidade": r["Quantidade"],
                "Valor dos Processos": r["Valor dos Processos"],
                "Valor Serviço": r["Valor Serviço"],
                "Média Mensal 2025": r["Média Mensal 2025"],
                "Tendência Linear": r["Tendência Linear"],
            }
            for r in nf_historico["ano_atual"]
        ]
        write_table(ws, next_row, 1, ["Ano", "Mês", "Período", "Quantidade", "Valor dos Processos", "Valor Serviço", "Média Mensal 2025", "Tendência Linear"], rows_current, "NFHistAtual")
        auto_width(ws)

    ws = wb.create_sheet("Deferimentos")
    write_title(ws, "Deferimentos", subtitle)
    next_row = write_table(ws, 4, 1, ["Empresa", "Quantidade", "Valor Deferido", "% do Valor"], deferimentos["por_empresa"], "DefEmpresa")
    next_row = write_table(ws, next_row, 1, ["Empresa", "Parceiro", "Município", "Estado", "Quantidade", "Valor Deferido"], deferimentos["por_municipio"], "DefMunicipio")
    write_table(ws, next_row, 1, ["Empresa", "Parceiro", "Município", "Estado", "Data Deferimento", "Valor Deferido", "Reclamação", "Tese", "Linha Original", "Arquivo"], deferimentos["records"], "DefDetalhe")
    auto_width(ws)

    if deferimentos_pendentes is not None:
        ws = wb.create_sheet("Deferimentos Pendentes")
        write_title(ws, "Deferimentos com pagamento pendente", f"Base: deferimentos de 01/01/2025 a {end.strftime('%d/%m/%Y')}")
        headers_summary = ["Situação do Pagamento", "Quantidade", "Valor Deferido", "Valor Pago Até Agora", "Saldo Pendente"]
        next_row = write_table(ws, 4, 1, headers_summary, deferimentos_pendentes["por_situacao"], "DefPendSituacao")
        next_row = write_table(ws, next_row, 1, ["Empresa", "Quantidade", "Valor Deferido", "Valor Pago Até Agora", "Saldo Pendente"], deferimentos_pendentes["por_empresa"], "DefPendEmpresa")
        write_table(ws, next_row, 1, ["Empresa", "Parceiro", "Município", "Estado", "Data Deferimento", "Reclamação", "Tese", "Situação do Pagamento", "Valor Deferido", "Valor Pago Até Agora", "Saldo Pendente", "% Pago", "Linha Original", "Arquivo"], deferimentos_pendentes["records"], "DefPendDetalhe")
        auto_width(ws)

    ws = wb.create_sheet("Inconsistências")
    write_title(ws, "Observações e inconsistências", subtitle)
    write_table(ws, 4, 1, ["Origem", "Arquivo", "Aba", "Linha", "Tipo", "Detalhe"], inconsistencias, "Inconsistencias")
    auto_width(ws)

    if output_stem is None:
        output_stem = f"apuracao_mensal_{month:02d}_{year}"
    output = OUTPUTS / f"{output_stem}.xlsx"
    wb.save(output)
    return output


def main():
    args = sys.argv[1:]
    if not args:
        args = ["--periodo"]

    if args and args[0].lower() in {"--periodo", "periodo", "--intervalo", "intervalo"}:
        if len(args) >= 3:
            start = parse_interval_date(args[1])
            end = parse_interval_date(args[2])
        else:
            start = parse_interval_date(input("Informe a data inicial (DD/MM/AAAA): "))
            end = parse_interval_date(input("Informe a data final (DD/MM/AAAA): "))
        if end < start:
            raise ValueError("A data final não pode ser menor que a data inicial.")

        ingressos = apurar_ingressos(start, end)
        tarefas = apurar_tarefas(start, end)
        nf = apurar_nf(start, end)
        deferimentos = apurar_deferimentos(start, end)
        deferimentos_pendentes = apurar_deferimentos_pendentes(end)
        inconsistencias = (
            ingressos["inconsistencias"]
            + tarefas["inconsistencias"]
            + nf["inconsistencias"]
            + deferimentos["inconsistencias"]
            + deferimentos_pendentes["inconsistencias"]
        )
        output_stem = f"apuracao_periodo_{start.strftime('%Y%m%d')}_a_{end.strftime('%Y%m%d')}"
        output = build_workbook(None, None, start, end, ingressos, tarefas, nf, deferimentos, inconsistencias, deferimentos_pendentes=deferimentos_pendentes, output_stem=output_stem, title="Apuração por período consolidada")
        detail = {
            "periodo": {
                "mes": None,
                "ano": None,
                "inicio": start.isoformat(),
                "fim": end.isoformat(),
                "rotulo": f"{start.strftime('%d/%m/%Y')} a {end.strftime('%d/%m/%Y')}",
            },
            "ingressos": ingressos,
            "tarefas": tarefas,
            "nf": nf,
            "nf_historico": None,
            "deferimentos": deferimentos,
            "deferimentos_pendentes": deferimentos_pendentes,
            "comparativo": None,
            "inconsistencias": inconsistencias,
        }
        detail_path = OUTPUTS / f"{output_stem}_detalhado.json"
        detail_path.write_text(json.dumps(detail, ensure_ascii=False, indent=2), encoding="utf-8")
        summary = {
            "arquivo": str(output),
            "tipo_apuracao": "intervalo",
            "periodo": f"{start.isoformat()} a {end.isoformat()}",
            "ingressos_total": ingressos["totais"]["Total"],
            "atividades_total": tarefas["total"],
            "nf_anderson_valor": nf["totais"]["Valor Anderson"],
            "deferimentos_quantidade": deferimentos["totais"]["Quantidade"],
            "deferimentos_valor": deferimentos["totais"]["Valor Deferido"],
            "deferimentos_saldo_pendente": deferimentos_pendentes["totais"]["Saldo Pendente"],
            "inconsistencias": len(inconsistencias),
        }
        json_path = OUTPUTS / f"{output_stem}.json"
        json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return

    raise ValueError("Esta versão exportável executa apenas apuração por período. Use: python apuracao_periodo.py --periodo DD/MM/AAAA DD/MM/AAAA")

    if args and args[0].lower() in {"--mensal", "mensal"}:
        arg = args[1] if len(args) > 1 else input("Informe o mês/ano (MM/AAAA): ")
    else:
        arg = sys.argv[1] if len(sys.argv) > 1 else input("Informe o mês/ano (MM/AAAA): ")
    month, year, start, end = parse_period(arg)
    prev_month, prev_year, prev_start, prev_end = previous_period(start)
    ingressos = apurar_ingressos(start, end)
    tarefas = apurar_tarefas(start, end)
    nf = apurar_nf(start, end)
    nf_historico = apurar_nf_historico(year, month)
    deferimentos = apurar_deferimentos(start, end)
    deferimentos_pendentes = apurar_deferimentos_pendentes(end)
    prev_data = {
        "ingressos": apurar_ingressos(prev_start, prev_end),
        "tarefas": apurar_tarefas(prev_start, prev_end),
        "nf": apurar_nf(prev_start, prev_end),
        "deferimentos": apurar_deferimentos(prev_start, prev_end),
    }
    current_data = {
        "ingressos": ingressos,
        "tarefas": tarefas,
        "nf": nf,
        "nf_historico": nf_historico,
        "deferimentos": deferimentos,
    }
    comparativo = build_comparativo(current_data, prev_data)
    inconsistencias = (
        ingressos["inconsistencias"]
        + tarefas["inconsistencias"]
        + nf["inconsistencias"]
        + deferimentos["inconsistencias"]
        + deferimentos_pendentes["inconsistencias"]
    )
    output = build_workbook(month, year, start, end, ingressos, tarefas, nf, deferimentos, inconsistencias, comparativo, prev_start, prev_end, nf_historico, deferimentos_pendentes=deferimentos_pendentes)
    detail = {
        "periodo": {
            "mes": month,
            "ano": year,
            "inicio": start.isoformat(),
            "fim": end.isoformat(),
            "rotulo": f"{month:02d}/{year}",
        },
        "periodo_anterior": {
            "mes": prev_month,
            "ano": prev_year,
            "inicio": prev_start.isoformat(),
            "fim": prev_end.isoformat(),
            "rotulo": f"{prev_month:02d}/{prev_year}",
        },
        "ingressos": ingressos,
        "tarefas": tarefas,
        "nf": nf,
        "nf_historico": nf_historico,
        "deferimentos": deferimentos,
        "deferimentos_pendentes": deferimentos_pendentes,
        "comparativo": comparativo,
        "inconsistencias": inconsistencias,
    }
    detail_path = OUTPUTS / f"apuracao_mensal_{month:02d}_{year}_detalhado.json"
    detail_path.write_text(json.dumps(detail, ensure_ascii=False, indent=2), encoding="utf-8")

    ppt_output = None
    ppt_error = None
    node = Path(os.environ.get("USERPROFILE", "")) / ".cache" / "codex-runtimes" / "codex-primary-runtime" / "dependencies" / "node" / "bin" / "node.exe"
    ppt_script = ROOT / "work" / "build_ppt_mensal.js"
    if node.exists() and ppt_script.exists():
        env = os.environ.copy()
        deps = Path(os.environ.get("USERPROFILE", "")) / ".cache" / "codex-runtimes" / "codex-primary-runtime" / "dependencies" / "node" / "node_modules" / ".pnpm"
        env["NODE_PATH"] = ";".join([
            str(deps / "pptxgenjs@4.0.1" / "node_modules"),
            str(deps / "jszip@3.10.1" / "node_modules"),
        ])
        result = subprocess.run(
            [str(node), str(ppt_script), f"{month:02d}/{year}"],
            cwd=str(ROOT),
            env=env,
            capture_output=True,
            text=True,
        )
        if result.returncode == 0:
            ppt_output = result.stdout.strip().splitlines()[-1] if result.stdout.strip() else None
        else:
            ppt_error = (result.stderr or result.stdout or "Erro desconhecido ao gerar PPT").strip()
    else:
        ppt_error = "Node ou script do PowerPoint não encontrado."

    summary = {
        "arquivo": str(output),
        "powerpoint": ppt_output,
        "powerpoint_erro": ppt_error,
        "periodo": f"{start.isoformat()} a {end.isoformat()}",
        "periodo_anterior": f"{prev_start.isoformat()} a {prev_end.isoformat()}",
        "ingressos_total": ingressos["totais"]["Total"],
        "atividades_total": tarefas["total"],
        "nf_anderson_valor": nf["totais"]["Valor Anderson"],
        "deferimentos_quantidade": deferimentos["totais"]["Quantidade"],
        "deferimentos_valor": deferimentos["totais"]["Valor Deferido"],
        "deferimentos_saldo_pendente": deferimentos_pendentes["totais"]["Saldo Pendente"],
        "inconsistencias": len(inconsistencias),
    }
    json_path = OUTPUTS / f"apuracao_mensal_{month:02d}_{year}.json"
    json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
