import re
import sys
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from texter_utils import aba_historico_consumo, aba_info_geral, carregar_arquivo, historico
from Texter_format_functions.texter_format_energisa import format_energisa


if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent.parent

PATH_POPPLER = BASE_DIR / "Faturas_Poppler"
PATH_ANALISE = BASE_DIR / "Faturas_Analaiser"

MESES = {
    "JAN": 1,
    "FEV": 2,
    "MAR": 3,
    "ABR": 4,
    "MAI": 5,
    "JUN": 6,
    "JUL": 7,
    "AGO": 8,
    "SET": 9,
    "OUT": 10,
    "NOV": 11,
    "DEZ": 12,
}


def limpar_estado_processamento():
    aba_info_geral.clear()
    aba_historico_consumo.clear()
    historico.clear()


def parse_data(data_str):
    mes, ano = data_str.split("/")
    mes = mes.strip().upper()
    ano = re.sub(r"\D", "", ano)

    if len(ano) == 0:
        raise ValueError(f"Ano invalido em: {data_str}")

    return (2000 + int(ano), MESES[mes])


def transformar(matriz_consumo):
    datas = set()
    for linha in matriz_consumo:
        for data, _ in linha[1:]:
            datas.add(data)

    datas_ordenadas = sorted(datas, key=parse_data)

    resultado = []
    resultado.append(["DATA"] + datas_ordenadas)

    for linha in matriz_consumo:
        unidade = linha[0]
        mapa = {data: valor for data, valor in linha[1:]}

        nova_linha = [unidade]
        for data in datas_ordenadas:
            nova_linha.append(mapa.get(data, "UNK"))

        resultado.append(nova_linha)

    return resultado


def formatar_planilha_visual(arquivo):
    wb = load_workbook(arquivo)

    fonte_padrao = Font(name="Verdana", size=8)
    fonte_cabecalho = Font(name="Verdana", size=8, bold=True)
    alinhamento = Alignment(horizontal="center", vertical="center")
    borda_fina = Side(style="thin")
    borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    fill_claro = PatternFill(fill_type="solid", fgColor="FFFFFF")
    fill_escuro = PatternFill(fill_type="solid", fgColor="DCE6F1")

    def formatar_aba(ws):
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            largura_max = 0
            for cell in col:
                if cell.value is not None:
                    largura_max = max(largura_max, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(largura_max + 2, 8)

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            eh_cabecalho = row_idx == 1
            for cell in row:
                cell.font = fonte_cabecalho if eh_cabecalho else fonte_padrao
                cell.alignment = alinhamento
                cell.border = borda
                if eh_cabecalho:
                    cell.fill = fill_claro
                else:
                    cell.fill = fill_escuro if row_idx % 2 == 0 else fill_claro

    if "INFORMACAO GERAL" in wb.sheetnames:
        formatar_aba(wb["INFORMACAO GERAL"])
    if "HISTORICO DE CONSUMO" in wb.sheetnames:
        formatar_aba(wb["HISTORICO DE CONSUMO"])

    wb.save(arquivo)


def gerar_planilha_analise(municipio_dir):
    limpar_estado_processamento()

    arquivos_txt = sorted([f for f in municipio_dir.iterdir() if f.is_file() and f.suffix.lower() == ".txt"])
    if not arquivos_txt:
        print(f"[AVISO] Sem arquivos txt em {municipio_dir.name}.")
        return None

    erros = []
    for arquivo in arquivos_txt:
        try:
            conteudo_bruto = carregar_arquivo(arquivo)
            format_energisa(conteudo_bruto, arquivo.name)
        except Exception as exc:
            erros.append((arquivo.name, str(exc)))

    colunas_base_info_geral = [
        "UNIDADE",
        "FORNECIMENTO",
        "NIVEL DE TENSAO",
        "CODIGO",
        "CLASSIFICACAO",
        "DESTINO",
        "ENDERECO",
        "MEDIDOR",
    ]

    max_colunas_info_geral = max((len(linha) for linha in aba_info_geral), default=0)
    if max_colunas_info_geral > 0:
        cabecalho_info_geral = []
        for i in range(max_colunas_info_geral):
            if i < len(colunas_base_info_geral):
                cabecalho_info_geral.append(colunas_base_info_geral[i])
            elif i == max_colunas_info_geral - 1:
                cabecalho_info_geral.append("COMPLEMENTO")
            else:
                cabecalho_info_geral.append(f"CAMPO_{i + 1}")

        info_geral = [cabecalho_info_geral] + aba_info_geral
    else:
        info_geral = [["SEM DADOS"]]

    grupos_historico_consumo = defaultdict(list)
    for linha in aba_historico_consumo:
        if not linha:
            continue
        chave = linha[0]
        grupos_historico_consumo[chave].extend(linha[1:])

    historico_consumo_agrupado = []
    for chave, valores in grupos_historico_consumo.items():
        historico_consumo_agrupado.append([chave] + valores)

    if historico_consumo_agrupado:
        historico_consumo_tratado = transformar(historico_consumo_agrupado)
    else:
        historico_consumo_tratado = [["SEM DADOS"]]

    df_info_geral = pd.DataFrame(info_geral)
    df_historico_consumo = pd.DataFrame(historico_consumo_tratado)

    output_dir_name = municipio_dir.name.replace("Poppler", "Analaiser")
    output_dir = PATH_ANALISE / output_dir_name
    output_dir.mkdir(parents=True, exist_ok=True)

    nome_planilha = output_dir_name + ".xlsx"
    arquivo_saida = output_dir / nome_planilha

    with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
        df_info_geral.to_excel(writer, sheet_name="INFORMACAO GERAL", index=False, header=False)
        df_historico_consumo.to_excel(writer, sheet_name="HISTORICO DE CONSUMO", index=False, header=False)

    formatar_planilha_visual(arquivo_saida)

    print(f"[OK] Planilha gerada: {arquivo_saida}")
    if erros:
        log_erros = output_dir / "erros_processamento.txt"
        with open(log_erros, "w", encoding="utf-8") as f:
            for nome_arquivo, erro in erros:
                f.write(f"{nome_arquivo}\t{erro}\n")

        print(f"[AVISO] {len(erros)} arquivo(s) falharam em {municipio_dir.name}.")
        print(f"[AVISO] Log completo: {log_erros}")

        limite_preview = 20
        for nome_arquivo, erro in erros[:limite_preview]:
            print(f"  - {nome_arquivo}: {erro}")
        if len(erros) > limite_preview:
            print(f"  ... e mais {len(erros) - limite_preview} falha(s).")

    return arquivo_saida


def listar_pastas_municipios_poppler():
    if not PATH_POPPLER.exists():
        return []

    pastas = []
    for pasta in PATH_POPPLER.iterdir():
        if not pasta.is_dir():
            continue
        nome_upper = pasta.name.upper()
        if not nome_upper.endswith("_POPPLER"):
            continue
        if "LAYOUT" in nome_upper:
            continue
        pastas.append(pasta)

    return sorted(pastas, key=lambda p: p.name)


def main():
    PATH_ANALISE.mkdir(parents=True, exist_ok=True)

    pastas_municipios = listar_pastas_municipios_poppler()
    if not pastas_municipios:
        print("Nenhuma pasta de municipio Poppler encontrada.")
        return

    print("Pastas de municipios encontradas:")
    for pasta in pastas_municipios:
        print(f" - {pasta.name}")

    print("\nGerando planilhas de analise Energisa...")
    for pasta in pastas_municipios:
        gerar_planilha_analise(pasta)

    print("\nConcluido.")


if __name__ == "__main__":
    main()
