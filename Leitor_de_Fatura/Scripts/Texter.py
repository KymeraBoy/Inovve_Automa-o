# ============================================================== #
#     BIBLIOTECAS
# ============================================================== #

import sys
import shutil
import csv
from pathlib            import Path
from openpyxl           import Workbook
from openpyxl.styles    import Font, Alignment

from Texter_format_functions.texter_format_enel         import format_enel
from Texter_format_functions.texter_format_energisa     import format_energisa
from Texter_format_functions.texter_format_neoenergia   import format_neoenergia

# ============================================================== #
# CONFIGURACOES (Serão sobrescritas dinamicamente pela GUI)
# ============================================================== #

PATH_POPPLER_PASTE      = Path(".")
PATH_TEXTER_PASTE       = Path(".")
PATH_ANALAISER_PASTE    = Path(".")

# Configuração central das abas do relatório.
# Para adicionar uma nova aba, basta incluir um novo dicionário na lista.
ABAS_RELATORIO_CONFIG = [
    {"nome_aba": "Classificação", "campo": "Classificação", "valor_padrao": ""},
    {"nome_aba": "Consumo_Medido", "campo": "Consumo Medido", "valor_padrao": 0.0},
    {"nome_aba": "Consumo_Faturado", "campo": "Consumo Faturado", "valor_padrao": 0.0},
    {"nome_aba": "Fornecimento", "campo": "Fornecimento", "valor_padrao": ""},
    {"nome_aba": "Cliente", "campo": "Cliente", "valor_padrao": ""},
    {"nome_aba": "Endereço", "campo": "Endereço", "valor_padrao": ""},
]

MESES_MAP = {
    'JAN': 1, 'FEV': 2, 'MAR': 3, 'ABR': 4, 'MAI': 5, 'JUN': 6, 'JUL': 7, 'AGO': 8, 'SET': 9, 'OUT': 10, 'NOV': 11, 'DEZ': 12
}

CONCESSIONÁRIA_MAP = {"NEOENERGIA": 1, "ENEL": 2, "ENERGISA": 3}
# ============================================================== #
# FUNCOES
# ============================================================== #


def limpar_pasta(caminho_pasta: Path) -> None:
    pasta = Path(caminho_pasta)
    if not pasta.exists():
        raise FileNotFoundError(f"A pasta '{pasta}' não existe.")
    if not pasta.is_dir():
        raise NotADirectoryError(f"'{pasta}' não é uma pasta.")
    for item in pasta.iterdir():
        if item.is_file() or item.is_symlink():
            item.unlink()
        elif item.is_dir():
            shutil.rmtree(item)

def selecionar_subapasta(PATH: Path, municipio_name: str) -> Path:
    subfolders = [f.name for f in PATH.iterdir() if f.is_dir()]
    for folder_name in subfolders:
        if folder_name == municipio_name:
            return PATH / folder_name
    raise ValueError(f"Município '{municipio_name}' não encontrado em '{PATH}'.")

def chave_ordenacao_mes(mes_ano):
    try:
        if '/' in str(mes_ano):
            partes = str(mes_ano).split('/')
            if partes[0].isdigit():
                return (int(partes[1]), int(partes[0]))
            else:
                return (int(partes[1]), MESES_MAP.get(partes[0].upper(), 0))
    except:
        pass
    return (9999, 0)

def gerar_base_matriz_vazia(lista_faturas_tagueadas):
    ucs_unicas = set()
    meses_unicos = set()

    for fatura in lista_faturas_tagueadas:
        uc = fatura.get("Unidade Consumidora")
        mes = fatura.get("Mês de referência")
        if uc:
            ucs_unicas.add(str(uc).strip())
        if mes:
            meses_unicos.add(str(mes).strip())

    lista_ucs_ordenada = sorted(list(ucs_unicas))
    lista_meses_ordenada = sorted(list(meses_unicos), key=chave_ordenacao_mes)

    matriz_base = []
    cabecalho = ["Unidade Consumidora"] + lista_meses_ordenada
    matriz_base.append(cabecalho)

    for uc in lista_ucs_ordenada:
        linha_vazia = [uc] + [None] * len(lista_meses_ordenada)
        matriz_base.append(linha_vazia)

    return matriz_base

def preencher_matriz_com_tag(matriz_base, lista_faturas_tagueadas, tag_valor):
    import copy
    matriz_preenchida = copy.deepcopy(matriz_base)
    cabecalho_meses = matriz_preenchida[0]

    mapa_busca = {}
    for fatura in lista_faturas_tagueadas:
        uc = str(fatura.get("Unidade Consumidora", "")).strip()
        mes = str(fatura.get("Mês de referência", "")).strip()
        valor = fatura.get(tag_valor, 0.0)
        if uc and mes:
            if uc not in mapa_busca:
                mapa_busca[uc] = {}
            mapa_busca[uc][mes] = valor

    for i in range(1, len(matriz_preenchida)):
        linha = matriz_preenchida[i]
        uc_linha = str(linha[0]).strip()
        for j in range(1, len(linha)):
            mes_coluna = str(cabecalho_meses[j]).strip()
            if uc_linha in mapa_busca and mes_coluna in mapa_busca[uc_linha]:
                linha[j] = mapa_busca[uc_linha][mes_coluna]
            else:
                linha[j] = 0.0

    return matriz_preenchida

def gerar_matriz_resumo_mais_recente(lista_faturas_tagueadas, campos_resumo):
    """
    Gera uma matriz-resumo com o valor mais recente de cada campo por UC.

    A coluna 1 permanece "Unidade Consumidora" e as demais colunas são
    os nomes das abas/campos informados em campos_resumo.
    """
    por_uc = {}

    for fatura in lista_faturas_tagueadas:
        uc = str(fatura.get("Unidade Consumidora", "")).strip()
        mes = str(fatura.get("Mês de referência", "")).strip()
        if not uc:
            continue

        chave_mes = chave_ordenacao_mes(mes)
        atual = por_uc.get(uc)
        if (atual is None) or (chave_mes > atual["chave_mes"]):
            por_uc[uc] = {
                "chave_mes": chave_mes,
                "dados": {nome_coluna: fatura.get(campo, valor_padrao) for nome_coluna, campo, valor_padrao in campos_resumo},
            }

    matriz = [["Unidade Consumidora"] + [nome_coluna for nome_coluna, _, _ in campos_resumo]]

    for uc in sorted(por_uc.keys()):
        dados = por_uc[uc]["dados"]
        linha = [uc] + [dados.get(nome_coluna, "") for nome_coluna, _, _ in campos_resumo]
        matriz.append(linha)

    return matriz

def gerar_abas_detalhadas_por_config(matriz_base, lista_faturas_tagueadas, abas_config):
    """
    Gera as matrizes detalhadas (por mês) a partir da configuração de abas.
    """
    abas = {}
    for cfg in abas_config:
        nome_aba = cfg["nome_aba"]
        campo = cfg["campo"]
        abas[nome_aba] = preencher_matriz_com_tag(matriz_base, lista_faturas_tagueadas, campo)
    return abas

def gerar_campos_resumo_por_config(abas_config):
    """
    Converte a configuração de abas para o formato esperado pela matriz de resumo.
    """
    return [
        (cfg["nome_aba"], cfg["campo"], cfg.get("valor_padrao", 0.0))
        for cfg in abas_config
    ]

def exportar_matrizes_para_xlsx(dicionario_abas, pasta_destino, nome_arquivo="Relatorio_Consolidado.xlsx"):
    diretorio = Path(pasta_destino)
    diretorio.mkdir(parents=True, exist_ok=True)
    caminho_final = diretorio / nome_arquivo

    wb = Workbook()
    primeira_aba = True

    for nome_aba, matriz in dicionario_abas.items():
        if primeira_aba:
            ws = wb.active
            ws.title = nome_aba
            primeira_aba = False
        else:
            ws = wb.create_sheet(title=nome_aba)

        for linha in matriz:
            linha_tratada = [item if item is not None else "" for item in linha]
            ws.append(linha_tratada)

    fonte_v = Font(name='Verdana', size=8)
    align_c = Alignment(horizontal='center', vertical='center')

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = fonte_v
                cell.alignment = align_c
                if isinstance(cell.value, (int, float)) and cell.row > 1:
                    cell.number_format = '#,##0.00'

        for col in sheet.columns:
            max_l = max([len(str(c.value)) for c in col if c.value] + [12])
            sheet.column_dimensions[col[0].column_letter].width = max_l + 4

    wb.save(caminho_final)

# ============================================================== #
# ORQUESTRADOR
# ============================================================== #

def texter_orchestrator(municipio_name: str, concessionaria_name: str, progress_callback=None):

    PATH_POPPLER_PASTE.mkdir(parents=True, exist_ok=True)  # Checar - Existência - Pasta Poppler
    PATH_TEXTER_PASTE.mkdir(parents=True, exist_ok=True)    # Checar - Existência - Pasta Texter
    PATH_ANALAISER_PASTE.mkdir(parents=True, exist_ok=True) # Checar - Existência - Pasta Analaiser

    pop_dir = PATH_POPPLER_PASTE / f"{municipio_name}_Poppler"  # Atribui - Endereço - Pasta Poppler do Município
    txt_dir = PATH_TEXTER_PASTE / f"{municipio_name}_Texter"    # Atribui - Endereço - Pasta Texter do Município
    txt_dir.mkdir(parents=True, exist_ok=True)                  # Checar - Existência - Pasta Texter do Município
    pop_dir.mkdir(parents=True, exist_ok=True)                  # Checar - Existência
    print(pop_dir)
    limpar_pasta(txt_dir)                                       # Limpar - Pasta Texter do Município

    formatacao = CONCESSIONÁRIA_MAP.get(concessionaria_name.upper())
    if formatacao is None:
        raise ValueError(f"Concessionária '{concessionaria_name}' não reconhecida para o Texter.")
    
    files = sorted([f.name for f in pop_dir.iterdir() if f.is_file() and f.suffix.lower() == ".txt"])
    matriz = []
    total_files = len(files)

    for idx, file_name in enumerate(files):
        if progress_callback:
            progress_callback(idx + 1, total_files, f"Texter: Processando {file_name} ({idx + 1}/{total_files})...")
        input_path = pop_dir / file_name
        ind_data = None
        if formatacao == 1:
            ind_data = format_neoenergia(input_path, file_name)
            
        
        matriz.append(ind_data)  

      

    matriz_base     = gerar_base_matriz_vazia(matriz)
    abas_detalhadas = gerar_abas_detalhadas_por_config(matriz_base, matriz, ABAS_RELATORIO_CONFIG)
    matriz_resumo   = gerar_matriz_resumo_mais_recente(matriz, gerar_campos_resumo_por_config(ABAS_RELATORIO_CONFIG))

    abas_exportacao = {"Resumo": matriz_resumo}
    abas_exportacao.update(abas_detalhadas)
   
    exportar_matrizes_para_xlsx(
        abas_exportacao,
        PATH_ANALAISER_PASTE,
        nome_arquivo=f"Relatorio_Consolidado_{municipio_name}.xlsx"
    )

    print("\nFluxo Texter finalizado.")

if __name__ == "__main__":
    print("Este script não deve ser executado diretamente. Use a GUI.")